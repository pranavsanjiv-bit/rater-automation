"""
excel_generator.py
Builds a downloadable .xlsx workbook from the JSON definition produced by the
Streamlit dimension template generator.
"""

import io
import re
import itertools

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter


# ── Colours ───────────────────────────────────────────────────────────────────

CLR_HEADER_BG   = "1F3864"   # dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_INDEX_BG    = "D9E1F2"
CLR_INDEX_ALT   = "EEF2F9"
CLR_SUBHDR_BG   = "4472C4"
CLR_SUBHDR_FG   = "FFFFFF"
CLR_AXIS_BG     = "F2F7FF"
CLR_AXIS_FG     = "1F3864"
CLR_BORDER      = "BDD7EE"
CLR_LEGEND_BG   = "FCE8D6"
CLR_LEGEND_ALT  = "FDF3E7"
CLR_LEGEND_HDR  = "C66A1E"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _font(bold=False, color="000000", size=10, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _thin_border():
    s = Side(style="thin", color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Slug utilities ────────────────────────────────────────────────────────────

_INVALID_CHARS = re.compile(r"[/\\?*\[\]:]")

def _safe_sheet_name(name: str, existing: set) -> str:
    """Truncate to 31 chars, strip invalid chars, guarantee uniqueness."""
    name = _INVALID_CHARS.sub("_", name)
    if len(name) > 31:
        name = name[:31]
    candidate = name
    suffix = 2
    while candidate in existing:
        tag = f"_{suffix}"
        candidate = name[: 31 - len(tag)] + tag
        suffix += 1
    existing.add(candidate)
    return candidate


# ── Cartesian product of parsed_dimensions ───────────────────────────────────

def _build_combinations(parsed_dimensions: dict) -> list[dict]:
    """
    Returns list of dicts like:
      [{"ct": ("reducing", "Reducing"), "si": ("lt_20l", "Less than 20L"), ...}, ...]
    i.e. each entry maps dim_key → (slug_key, display_label)
    """
    if not parsed_dimensions:
        return [{}]

    keys = list(parsed_dimensions.keys())
    value_lists = []
    for k in keys:
        dim_vals = parsed_dimensions[k]
        if not isinstance(dim_vals, dict):
            raise ValueError(
                f"parsed_dimensions['{k}'] must be a dict mapping slug→label, "
                f"got {type(dim_vals).__name__}: {dim_vals!r}"
            )
        if not dim_vals:
            continue
        value_lists.append([(slug, label) for slug, label in dim_vals.items()])

    if not value_lists:
        return [{}]

    combos = []
    for product in itertools.product(*value_lists):
        combo = {keys[i]: product[i] for i in range(len(keys))}
        combos.append(combo)
    return combos


def _sequenced_sheet_name(index: int, existing: set) -> str:
    """Return a sequenced sheet name like Sheet001, Sheet002, …"""
    name = f"Sheet{index:03d}"
    candidate = name
    suffix = 2
    while candidate in existing:
        candidate = f"{name}_{suffix}"
        suffix += 1
    existing.add(candidate)
    return candidate


def _unlock_data_cells(ws, two_d: bool, num_axis_rows: int, num_axis_cols: int):
    """Unlock only the rate-entry cells; headers and axis labels stay locked."""
    locked = Protection(locked=True)
    unlocked = Protection(locked=False)

    if two_d:
        for row in ws.iter_rows(min_row=1, max_row=num_axis_rows + 1, min_col=1, max_col=num_axis_cols + 1):
            for cell in row:
                cell.protection = locked
        for ri in range(2, num_axis_rows + 2):
            for ci in range(2, num_axis_cols + 2):
                ws.cell(row=ri, column=ci).protection = unlocked
    else:
        for row in ws.iter_rows(min_row=1, max_row=num_axis_rows + 1, min_col=1, max_col=2):
            for cell in row:
                cell.protection = locked
        for ri in range(2, num_axis_rows + 2):
            ws.cell(row=ri, column=2).protection = unlocked

    ws.protection.sheet = True


# ── Axis expansion ────────────────────────────────────────────────────────────

def _expand_axis(axis_def: dict) -> list:
    """Return list of row/col header values for a given axis definition."""
    t = axis_def.get("type", "Range")
    if t == "Range":
        lo = int(axis_def.get("min", 0))
        hi = int(axis_def.get("max", 10))
        return list(range(lo, hi + 1))
    if t == "Enum":
        cfg = axis_def.get("config", [])
        if isinstance(cfg, list):
            return cfg
        return cfg.get("values", [])
    if t == "Comparison":
        cfg = axis_def.get("config", [])
        return [f"{c['op']} {c['val']}" for c in cfg]
    return []


# ── Data sheet builder ────────────────────────────────────────────────────────

def _write_data_sheet_2d(ws, row_axis: dict, col_axis: dict, combo: dict):
    """2D grid: row axis name at A1, column axis values across row 1, row axis
    values down column A. No banner row, no combo metadata row — A1 holds the
    row-axis name (col axis name is implied by the header row, but to keep
    A1 = the row axis label specifically per spec, we put 'row \\ col' there)."""
    row_vals = _expand_axis(row_axis)
    col_vals = _expand_axis(col_axis)

    row_name = row_axis.get("name", "Row")
    col_name = col_axis.get("name", "Col")

    # A1 = corner label, no banner / metadata row above it
    corner = ws.cell(row=1, column=1, value=f"{row_name} \\ {col_name}")
    corner.font = _font(bold=True, color=CLR_AXIS_FG)
    corner.fill = _fill(CLR_AXIS_BG)
    corner.alignment = _center()
    corner.border = _thin_border()

    # Column headers (row 1, starting col B)
    for ci, cv in enumerate(col_vals, start=2):
        cell = ws.cell(row=1, column=ci, value=cv)
        cell.font = _font(bold=True, color=CLR_SUBHDR_FG)
        cell.fill = _fill(CLR_SUBHDR_BG)
        cell.alignment = _center()
        cell.border = _thin_border()

    # Row headers + empty data cells
    for ri, rv in enumerate(row_vals, start=2):
        rh = ws.cell(row=ri, column=1, value=rv)
        rh.font = _font(bold=True, color=CLR_AXIS_FG)
        rh.fill = _fill(CLR_AXIS_BG)
        rh.alignment = _center()
        rh.border = _thin_border()

        for ci in range(2, len(col_vals) + 2):
            dc = ws.cell(row=ri, column=ci, value=None)
            dc.border = _thin_border()

    ws.column_dimensions["A"].width = max(len(str(row_name)) + 4, 12)
    for ci in range(2, len(col_vals) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 8

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = ws.cell(row=2, column=2)


def _write_data_sheet_1d(ws, col_axis: dict, combo: dict):
    """1D layout used when the row axis is disabled. Axis values run down
    column A starting at A2, data goes in column B. A1 holds the axis name."""
    col_vals = _expand_axis(col_axis)
    col_name = col_axis.get("name", "Col")

    header = ws.cell(row=1, column=1, value=col_name)
    header.font = _font(bold=True, color=CLR_AXIS_FG)
    header.fill = _fill(CLR_AXIS_BG)
    header.alignment = _center()
    header.border = _thin_border()

    data_header = ws.cell(row=1, column=2, value="Value")
    data_header.font = _font(bold=True, color=CLR_SUBHDR_FG)
    data_header.fill = _fill(CLR_SUBHDR_BG)
    data_header.alignment = _center()
    data_header.border = _thin_border()

    for ri, cv in enumerate(col_vals, start=2):
        ah = ws.cell(row=ri, column=1, value=cv)
        ah.font = _font(bold=True, color=CLR_AXIS_FG)
        ah.fill = _fill(CLR_AXIS_BG)
        ah.alignment = _center()
        ah.border = _thin_border()

        dc = ws.cell(row=ri, column=2, value=None)
        dc.border = _thin_border()

    ws.column_dimensions["A"].width = max(len(str(col_name)) + 4, 12)
    ws.column_dimensions["B"].width = 14
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = ws.cell(row=2, column=2)


# ── Index sheet builder ───────────────────────────────────────────────────────

def _write_dimension_legend(ws_idx, dimensions: dict, parsed_dimensions: dict, start_col: int) -> int:
    """
    Writes a 'Dimensions' legend with one column per dimension. Each column's
    sub-header is the dimension name, and its values are stacked below it.
    """
    slugs = list(parsed_dimensions.keys())
    names = list(dimensions.keys())
    dim_count = len(slugs)
    if dim_count == 0:
        return start_col

    end_col = start_col + dim_count - 1

    # Title row, merged across all dimension columns
    title_cell = ws_idx.cell(row=1, column=start_col, value="Dimensions")
    ws_idx.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    title_cell.font = _font(bold=True, color=CLR_HEADER_FG, size=12)
    title_cell.fill = _fill(CLR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Sub-header row: one cell per dimension name
    for ci, (slug, name) in enumerate(zip(slugs, names), start=start_col):
        h = ws_idx.cell(row=2, column=ci, value=name.upper())
        h.font = _font(bold=True, color="FFFFFF")
        h.fill = _fill(CLR_LEGEND_HDR)
        h.alignment = _center()
        h.border = _thin_border()

    # Values stacked below each dimension's column
    max_rows = 0
    for ci, slug in enumerate(slugs, start=start_col):
        values = list(parsed_dimensions[slug].values())
        max_rows = max(max_rows, len(values))
        for ri, val_label in enumerate(values, start=3):
            fill_color = CLR_LEGEND_BG if ri % 2 == 1 else CLR_LEGEND_ALT
            c = ws_idx.cell(row=ri, column=ci, value=val_label)
            c.fill = _fill(fill_color)
            c.border = _thin_border()

    for ci in range(start_col, end_col + 1):
        ws_idx.column_dimensions[get_column_letter(ci)].width = 24

    return end_col


def _write_index_sheet(ws_idx, sheet_entries: list[dict], parsed_dimensions: dict, dimensions: dict):
    """
    sheet_entries: list of {"sheet_name": str, "combo": {dim_key: (slug, label)}}
    """
    dim_keys = list(parsed_dimensions.keys())
    dim_names = list(dimensions.keys())  # Original display names

    headers = ["Sheet"] + dim_names

    # Title row
    ws_idx.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws_idx.cell(row=1, column=1, value="Template Index")
    title.font = _font(bold=True, color=CLR_HEADER_FG, size=12)
    title.fill = _fill(CLR_HEADER_BG)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws_idx.row_dimensions[1].height = 22

    # Header row
    for ci, h in enumerate(headers, start=1):
        c = ws_idx.cell(row=2, column=ci, value=h.upper())
        c.font = _font(bold=True, color=CLR_SUBHDR_FG)
        c.fill = _fill(CLR_SUBHDR_BG)
        c.alignment = _center()
        c.border = _thin_border()

    # Data rows
    for ri, entry in enumerate(sheet_entries, start=3):
        fill_color = CLR_INDEX_BG if ri % 2 == 1 else CLR_INDEX_ALT
        sn = entry["sheet_name"]
        combo = entry["combo"]

        c = ws_idx.cell(row=ri, column=1, value=sn)
        c.font = _font(bold=True)
        c.fill = _fill(fill_color)
        c.border = _thin_border()

        for di, dk in enumerate(dim_keys, start=2):
            if dk in combo:
                _, label = combo[dk]
            else:
                label = ""
            dc = ws_idx.cell(row=ri, column=di, value=label)
            dc.fill = _fill(fill_color)
            dc.border = _thin_border()

    ws_idx.column_dimensions["A"].width = 36
    for ci in range(2, len(headers) + 1):
        ws_idx.column_dimensions[get_column_letter(ci)].width = 20

    # ── Dimension legend, placed two columns to the right of the mapping table ─
    legend_start_col = len(headers) + 2
    _write_dimension_legend(ws_idx, dimensions, parsed_dimensions, legend_start_col)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel(data: dict) -> bytes:
    """
    Build the .xlsx workbook entirely in memory and return raw bytes.

    Parameters
    ----------
    data : dict
        {
          "definition": {
            "dimensions": {dim_name: {...}, ...},
            "axes": {
              "row": {"name":..., "type":..., ...}   # required, never None
              "col": {"name":..., "type":..., ...} | None   # optional
            }
          },
          "parsed_dimensions": { dim_slug: {val_slug: display_label, ...}, ... }
        }

        Row axis is always required. If axes.col is None, sheets are
        generated as a 1D list keyed on the row axis only (values down
        column A, data in column B).
    """
    definition = data.get("definition", {})
    dimensions = definition.get("dimensions", {})
    parsed_dimensions = data.get("parsed_dimensions", {})
    axes = definition.get("axes", {})

    row_axis = axes.get("row")
    col_axis = axes.get("col")

    if not row_axis:
        raise ValueError("A Row axis is required to generate the workbook.")

    two_d = col_axis is not None

    combos = _build_combinations(parsed_dimensions)

    wb = Workbook()

    ws_idx = wb.active
    ws_idx.title = "Index"

    existing_names: set = {"Index"}
    sheet_entries = []

    for idx, combo in enumerate(combos, start=1):
        sn = _sequenced_sheet_name(idx, existing_names)
        sheet_entries.append({"sheet_name": sn, "combo": combo})

    _write_index_sheet(ws_idx, sheet_entries, parsed_dimensions, dimensions)
    ws_idx.protection.sheet = True

    for entry in sheet_entries:
        ws = wb.create_sheet(title=entry["sheet_name"])
        if two_d:
            row_vals = _expand_axis(row_axis)
            col_vals = _expand_axis(col_axis)
            _write_data_sheet_2d(ws, row_axis, col_axis, entry["combo"])
            _unlock_data_cells(ws, True, len(row_vals), len(col_vals))
        else:
            axis_vals = _expand_axis(row_axis)
            _write_data_sheet_1d(ws, row_axis, entry["combo"])
            _unlock_data_cells(ws, False, len(axis_vals), 0)

    sheet_mapping = {}
    for entry in sheet_entries:
        combo = entry["combo"]
        sheet_mapping[entry["sheet_name"]] = {
            dim_key: slug_key for dim_key, (slug_key, _) in combo.items()
        }

    # ── Hidden _meta sheet ────────────────────────────────────────────────────
    # Stores the full JSON definition so the flatten endpoint can validate
    # the uploaded file against the original template.
    ws_meta = wb.create_sheet(title="_meta")
    ws_meta.sheet_state = "veryHidden"
    ws_meta.protection.sheet = True

    import json as _json
    meta_payload = {
        "template_id": data.get("template_id"),          # may be None if not yet saved
        "definition": definition,
        "parsed_dimensions": parsed_dimensions,
        "calculation": data.get("calculation", {}),
        "sheet_mapping": sheet_mapping,
    }
    ws_meta.cell(row=1, column=1, value=_json.dumps(meta_payload))
    wb.security.lockStructure = True
    # ─────────────────────────────────────────────────────────────────────────

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()