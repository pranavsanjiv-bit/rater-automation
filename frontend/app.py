# pyrefly: ignore [missing-import]
import streamlit as st
import pandas as pd  # <-- ADD THIS
import io         # <-- ADD THIS
# pyrefly: ignore [missing-import]
from code_editor import code_editor as st_code_editor
import re
import time
import requests
import json
from dim_parser import (
    get_dimension_slug,
    get_value_slug,
    operator_to_slug,
    _template_required_dimensions,
    _build_comparison_row,
    _overall_premium_status,
    _LOS_DIMS,
    _FE_DIMS,
    _ALL_DIMS,
    _DIM_CATEGORY,
    _EXCLUDED_FROM_FORMULA,
    _PREDEFINED_NAME_TO_SLUG,
    _PREDEFINED_SLUG_TO_NAME,
)
from excel_generator import generate_excel
from formula_eval import validate_formula, FormulaError
from concurrent.futures import ThreadPoolExecutor
import openpyxl

st.set_page_config(page_title="Offline rater automation", layout="wide")

st.markdown(
    """
    <style>
    /* ── Global Streamlit noise suppression ──────────────────────────────── */
    [data-testid="InputInstructions"] { display: none !important; }

    /* ── Formula bar ─────────────────────────────────────────────────────── */
    input[placeholder*="formula"] {
        font-family: monospace !important;
        font-size: 1rem !important;
        background-color: #1e1e2e !important;
        color: #cdd6f4 !important;
        border: 1px solid #45475a !important;
    }
    .formula-bar {
        font-family: monospace;
        font-size: 1rem;
        background: #1e1e2e;
        color: #cdd6f4;
        padding: 0 0.75rem;
        border-radius: 6px;
        height: 2.4rem;
        line-height: 2.4rem;
        border: 1px solid #45475a;
        word-break: break-all;
        box-sizing: border-box;
    }
    .formula-bar-empty {
        color: #585b70;
        font-style: italic;
    }

    /* ── All buttons: baseline height + flex centering ───────────────────── */
    div[data-testid="stButton"] > button {
        height: 2.4rem;
        padding-top: 0 !important;
        padding-bottom: 0 !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        line-height: 1 !important;
    }

    /* ── Card row columns: vertically center every cell ─────────────────── */
    div[data-testid="stColumn"] > div[data-testid="stVerticalBlock"] {
        display: flex !important;
        flex-direction: column !important;
        justify-content: center !important;
        min-height: 2.4rem;
    }

    /* ── Role pill badges ────────────────────────────────────────────────── */
    .pill {
        display: inline-flex;
        align-items: center;
        padding: 0.2rem 0.6rem;
        border-radius: 12px;
        font-size: 0.82rem;
        white-space: nowrap;
        line-height: 1;
    }
    .pill-green { background: #1a4731; color: #4ade80; }
    .pill-blue  { background: #172554; color: #60a5fa; }

    /* ── FV slug badge ───────────────────────────────────────────────────── */
    .fv-slug {
        font-family: monospace;
        font-size: 0.82rem;
        background: #2a2a3d;
        color: #a6e3a1;
        padding: 0.2rem 0.5rem;
        border-radius: 4px;
        display: inline-block;
        line-height: 1.4;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

OPERATORS = ["<", "<=", ">", ">="]
SERVER_URL = "http://localhost:5050"


# ── Session state init ────────────────────────────────────────────────────────

def init_state():
    defaults = {
        "dims": [],
        "form_open": False,
        "form": _empty_form(),
        "last_result": None,
        # formula builder
        "formula_tokens": [],
        "formula_editor_input": "",
        # formula variables
        "formula_variables": [],   # list of {name, slug, default_value}
        "_fv_add_open": False,
        "_fv_editing": None,       # int index of row being edited, or None
        # cross-tab state
        "last_template_id": None,
        "template_name": "",
        "product_code": "",
        "rounding_rule": "nearest",
        # Premium flow state tracking
        "session_token": None,
        "session_flow_type": None,
        # Dimension Library admin tab (kept separate from the Template
        # Builder's own "form"/"form_open" state above)
        "lib_form_open": False,
        "lib_form": _empty_library_form(),
        "lib_editing_id": None,
        "lib_editing_is_system": False,
        "lib_delete_confirm_id": None,
        "lib_message": None,   # (level, text) shown once, then cleared — survives st.rerun()
        "lib_search": "",
        # Bumped every time a library form is *opened* (Edit or Add clicked);
        # folded into that form's widget keys so reopening (even the same
        # row) always starts from freshly-loaded data instead of whatever
        # Streamlit remembers under a reused widget key.
        "lib_form_nonce": 0,
        # Bulk Premium Calculator: last computed result workbook, kept until
        # a new file/template is processed (so an unrelated rerun doesn't
        # need to recompute or lose the download).
        "bulk_result_bytes": None,
        "bulk_result_name": None,
        # Optional user-supplied payload overrides, keyed by flow_type
        # ("save_loan" / "save_lead" / "create_lead"). Empty/absent = use
        # the server's hardcoded default template for that flow.
        "custom_payloads": {},
        "calc_flow_type": {"single": None, "bulk": None}
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _empty_form():
    return {
        "category": "LOS",
        "name": "",
        "type": "Enum",
        "values": [],
        "min": 0,
        "max": 100,
        "comparisons": [],
        "_val_input_seq": 0,
        "_comp_val_input_seq": 0,
        # "existing" = picked via the Dimension Library dropdown (definition
        # copied verbatim, not editable here); "custom" = typed freehand and
        # configured like any dimension always has been. None = popup not
        # answered yet.
        "name_mode": None,
        "_library_dim_id": None,
    }


def _empty_library_form():
    return {
        "name": "",
        "type": "Enum",
        "values": [],
        "min": 0,
        "max": 100,
        "comparisons": [],
        "_val_input_seq": 0,
        "_comp_val_input_seq": 0,
    }


# ── Role helpers ──────────────────────────────────────────────────────────────

def _row_dim():
    return next((d for d in st.session_state.dims if d.get("role") == "row"), None)


def _col_dim():
    return next((d for d in st.session_state.dims if d.get("role") == "col"), None)


def _set_role(target_dim, role):
    if role == "col" and _row_dim() is None:
        st.warning("Set a Row axis before choosing a Column axis.", icon="⚠️")
        return
    for d in st.session_state.dims:
        if d.get("role") == role:
            d["role"] = None
    target_dim["role"] = role
    st.session_state.last_result = None


def _clear_role(target_dim):
    role = target_dim.get("role")
    target_dim["role"] = None
    if role == "row":
        col = _col_dim()
        if col is not None:
            col["role"] = None
    st.session_state.last_result = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def dim_summary(dim):
    if dim["type"] == "Enum":
        vals = dim["config"].get("values", [])
        return ", ".join(vals) if vals else "—"
    if dim["type"] == "Range":
        return f"{dim['config']['min']} → {dim['config']['max']}"
    if dim["type"] == "Comparison":
        comps = dim["config"].get("comparisons", [])
        return ", ".join(f"{c['op']} {c['val']}" for c in comps) if comps else "—"
    return "—"


def build_dim_from_form(form):
    if form["type"] == "Enum":
        config = {"values": list(form["values"])}
    elif form["type"] == "Range":
        config = {"min": form["min"], "max": form["max"]}
    else:
        config = {"comparisons": [dict(c) for c in form["comparisons"]]}
    return {"name": form["name"].strip(), "type": form["type"], "config": config, "role": None}


def _render_enum_inputs(form_key, form):
    st.markdown("**Values**")
    seq = form.get("_val_input_seq", 0)
    input_key = f"{form_key}_new_val_{seq}"

    c1, c2 = st.columns([4, 1])
    with c1:
        st.text_input(
            "New value", value="", key=input_key,
            label_visibility="collapsed", placeholder="Type a value and press add →",
        )
    with c2:
        if st.button("Add value", key=f"{form_key}_add_val"):
            v = st.session_state[input_key].strip()
            if v and v not in form["values"]:
                form["values"].append(v)
            form["_val_input_seq"] = seq + 1
            st.rerun()

    vals = form["values"]
    if vals:
        st.write("")
        to_delete = None
        for i, v in enumerate(vals):
            c1, c2 = st.columns([5, 1])
            with c1:
                st.markdown(f"- {v}")
            with c2:
                if st.button("✕", key=f"{form_key}_del_val_{i}"):
                    to_delete = i
        if to_delete is not None:
            form["values"].pop(to_delete)
            st.rerun()
    else:
        st.caption("No values yet.")


def _render_comparison_inputs(form_key, form):
    st.markdown("**Conditions**")
    comps = form["comparisons"]
    to_delete = None
    for i, c in enumerate(comps):
        cc1, cc2, cc3, _ = st.columns([1.5, 3, 0.8, 0.2])
        with cc1:
            comps[i]["op"] = st.selectbox(
                "Op", OPERATORS, index=OPERATORS.index(c["op"]),
                key=f"{form_key}_comp_op_{i}", label_visibility="collapsed",
            )
        with cc2:
            comps[i]["val"] = st.text_input(
                "Val", value=c["val"], key=f"{form_key}_comp_val_{i}",
                label_visibility="collapsed", placeholder="value",
            )
        with cc3:
            if st.button("Remove", key=f"{form_key}_del_comp_{i}"):
                to_delete = i

    if to_delete is not None:
        form["comparisons"].pop(to_delete)
        st.rerun()

    seq = form.get("_comp_val_input_seq", 0)
    new_val_key = f"{form_key}_new_comp_val_{seq}"

    c1, c2, c3 = st.columns([1.5, 3, 2])
    with c1:
        new_op = st.selectbox("New op", OPERATORS, key=f"{form_key}_new_op", label_visibility="collapsed")
    with c2:
        st.text_input(
            "New val", key=new_val_key,
            label_visibility="collapsed", placeholder="value",
        )
    with c3:
        if st.button("＋ Add condition", key=f"{form_key}_add_comp"):
            new_cv = st.session_state[new_val_key]
            form["comparisons"].append({"op": new_op, "val": new_cv.strip()})
            form["_comp_val_input_seq"] = seq + 1
            st.rerun()


# ── LOS / FE predefined dimension names ──────────────────────────────────────
# (imported from dim_parser.py — single source of truth shared with
# _template_required_dimensions(), so this list and the Bulk Upload
# Spreadsheet's required columns can never drift apart)


# ── Dimension Library: HTTP helpers ───────────────────────────────────────────
# Thin wrappers around /dimension-library, following the same pattern as
# _fetch_templates()/_fetch_template_detail() further below.

def _fetch_dimension_library() -> list[dict]:
    try:
        resp = requests.get(f"{SERVER_URL}/dimension-library", timeout=10)
        if resp.ok:
            return resp.json()
    except requests.exceptions.ConnectionError:
        pass
    return []


def _create_library_dimension(name: str, dim_type: str, config: dict) -> dict:
    try:
        resp = requests.post(
            f"{SERVER_URL}/dimension-library",
            json={"name": name, "type": dim_type, "config": config},
            timeout=10,
        )
        if resp.ok:
            return {"ok": True, "data": resp.json()}
        try:
            err = resp.json().get("error", resp.text)
        except ValueError:
            err = resp.text
        return {"ok": False, "error": err}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": "Could not reach the Flask server."}


def _update_library_dimension(dim_id: str, name: str, dim_type: str, config: dict) -> dict:
    try:
        resp = requests.put(
            f"{SERVER_URL}/dimension-library/{dim_id}",
            json={"name": name, "type": dim_type, "config": config},
            timeout=10,
        )
        if resp.ok:
            return {"ok": True, "data": resp.json()}
        try:
            err = resp.json().get("error", resp.text)
        except ValueError:
            err = resp.text
        return {"ok": False, "error": err}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": "Could not reach the Flask server."}


def _delete_library_dimension(dim_id: str, force: bool = False) -> dict:
    try:
        resp = requests.delete(
            f"{SERVER_URL}/dimension-library/{dim_id}",
            params={"force": "true"} if force else {},
            timeout=10,
        )
        if resp.ok:
            return {"ok": True, "data": resp.json()}
        try:
            data = resp.json()
        except ValueError:
            data = {"error": resp.text}
        return {"ok": False, "error": data.get("error", resp.text), "usage_count": data.get("usage_count")}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": "Could not reach the Flask server."}


# ── Section: unified dimension table + role assignment ────────────────────────

def _render_dim_rows():
    dims = st.session_state.dims
    row_dim = _row_dim()
    col_dim = _col_dim()

    lh1, lh3, lh4, lh5 = st.columns([3.5, 1.3, 3, 0.7])
    with lh1:
        st.caption("Dimension")
    with lh3:
        st.caption("Type")
    with lh4:
        st.caption("Axis Role")
    with lh5:
        st.caption("")

    for i, dim in enumerate(dims):
        with st.container(border=True):
            c_name, c_type, c_role, c_del = st.columns([3.5, 1.3, 3, 0.7])

            with c_name:
                st.markdown(f"**{dim['name']}**")

            with c_type:
                st.markdown(f"<small>{dim['type']}</small>", unsafe_allow_html=True)

            role = dim.get("role")
            with c_role:
                if role == "row":
                    rc1, rc2 = st.columns([3, 1])
                    with rc1:
                        st.markdown('<span class="pill pill-green">● Row</span>',
                                    unsafe_allow_html=True)
                    with rc2:
                        if st.button("✕", key=f"unassign_{i}", help="Unassign Row"):
                            _clear_role(dim)
                            st.rerun()
                elif role == "col":
                    rc1, rc2 = st.columns([3, 1])
                    with rc1:
                        st.markdown('<span class="pill pill-blue">● Column</span>',
                                    unsafe_allow_html=True)
                    with rc2:
                        if st.button("✕", key=f"unassign_{i}", help="Unassign Column"):
                            _clear_role(dim)
                            st.rerun()
                else:
                    can_be_row = row_dim is None
                    can_be_col = (row_dim is not None) and (dim is not row_dim) and (col_dim is None)
                    help_col_text = None
                    if row_dim is None:
                        help_col_text = "Set a Row axis first"
                    elif col_dim is not None:
                        help_col_text = "Clear the current Column axis first"
                    elif dim is row_dim:
                        help_col_text = "Already the Row axis"
                    rb1, rb2 = st.columns(2)
                    with rb1:
                        if st.button("→ Row", key=f"set_row_{i}",
                                     disabled=not can_be_row,
                                     help=None if can_be_row else "Clear the current Row axis first",
                                     use_container_width=True):
                            _set_role(dim, "row")
                            st.rerun()
                    with rb2:
                        if st.button("→ Col", key=f"set_col_{i}",
                                     disabled=not can_be_col,
                                     help=help_col_text,
                                     use_container_width=True):
                            _set_role(dim, "col")
                            st.rerun()

            with c_del:
                if st.button("🗑", key=f"del_dim_{i}", help="Remove dimension"):
                    st.session_state.dims.pop(i)
                    st.session_state.last_result = None
                    st.rerun()


def _dim_from_library_entry(entry: dict) -> dict:
    """Copy a Dimension Library entry's stored definition directly into a new
    template dimension. Used verbatim — the IC never edits or even reviews
    it during template creation (per business requirement)."""
    return {
        "name": entry["name"],
        "type": entry["type"],
        "config": dict(entry["config"]),
        "role": None,
        "category": "Library",
    }


@st.dialog("Add a dimension")
def _add_dimension_popup():
    st.write("How would you like to add this dimension?")
    choice = st.radio(
        "Add dimension mode",
        ["Choose Existing Dimension", "Create New Dimension"],
        key="_add_dim_popup_choice",
        label_visibility="collapsed",
    )
    st.write("")
    if st.button("Continue", key="_add_dim_popup_continue", type="primary", use_container_width=True):
        st.session_state.form = _empty_form()
        st.session_state.form["name_mode"] = (
            "existing" if choice == "Choose Existing Dimension" else "custom"
        )
        st.session_state.form_open = True
        st.rerun()


def render_dims_section():
    for k in ["do_confirm", "do_cancel"]:
        if k not in st.session_state:
            st.session_state[k] = False

    hdr_col, btn_col = st.columns([3, 2])
    with hdr_col:
        st.markdown("#### Dimensions")
        st.caption("Dimensions that exist in your rate table / Excel.")
    with btn_col:
        st.write("")
        if not st.session_state.form_open:
            if st.button("＋ Add Dimension", key="open_form", use_container_width=True):
                _add_dimension_popup()

    if st.session_state.dims:
        _render_dim_rows()
    else:
        st.caption("No dimensions yet.")

    if st.session_state.form_open:
        form = st.session_state.form
        already_added = {d["name"].strip().lower() for d in st.session_state.dims}
        can_confirm = False

        with st.container(border=True):
            if form["name_mode"] == "existing":
                st.markdown("##### New dimension · Choose Existing Dimension")
                library = _fetch_dimension_library()
                available = [d for d in library if d["name"].strip().lower() not in already_added]

                if not library:
                    st.warning(
                        "Could not load the Dimension Library (server unreachable, or it's empty).",
                        icon="⚠️",
                    )
                elif not available:
                    st.info("All library dimensions have already been added to this template.", icon="ℹ️")
                else:
                    names = [d["name"] for d in available]
                    default_idx = names.index(form["name"]) if form["name"] in names else 0
                    chosen_name = st.selectbox("Dimension", names, index=default_idx, key="form_name_existing")
                    chosen = next(d for d in available if d["name"] == chosen_name)
                    form["name"] = chosen["name"]
                    form["_library_dim_id"] = chosen["id"]
                    st.caption(
                        f"`{chosen['type']}` — this definition will be copied into the template as-is "
                        "and can't be edited here."
                    )
                    can_confirm = True

            else:  # "custom"
                st.markdown("##### New dimension · Create New Dimension")
                col_dim, col_type = st.columns(2)

                with col_dim:
                    form["name"] = st.text_input(
                        "Dimension Name", value=form["name"], key="form_name_custom",
                        placeholder="e.g. Occupation",
                    )
                    form["category"] = "Custom"

                with col_type:
                    form["type"] = st.selectbox(
                        "Type", ["Enum", "Range", "Comparison"],
                        index=["Enum", "Range", "Comparison"].index(form["type"]),
                        key="form_type",
                    )

                name_clean = form["name"].strip()
                name_conflict = bool(name_clean) and name_clean.lower() in already_added
                if name_conflict:
                    st.warning(
                        f"A dimension named '{name_clean}' is already on this template. "
                        "Choose a different name.", icon="⚠️",
                    )

                if form["type"] == "Enum":
                    _render_enum_inputs("form", form)
                elif form["type"] == "Range":
                    c1, c2 = st.columns(2)
                    with c1:
                        form["min"] = st.number_input("Min", value=int(form["min"]), step=1, key="form_min")
                    with c2:
                        form["max"] = st.number_input("Max", value=int(form["max"]), step=1, key="form_max")
                elif form["type"] == "Comparison":
                    _render_comparison_inputs("form", form)

                can_confirm = bool(name_clean) and not name_conflict

            st.write("")
            btn_col1, btn_col2, _ = st.columns([3, 2, 3])
            with btn_col1:
                if st.button("Add dimension", key="form_confirm", type="primary",
                             disabled=not can_confirm, use_container_width=True):
                    st.session_state.do_confirm = True
            with btn_col2:
                if st.button("Cancel", key="form_cancel", use_container_width=True):
                    st.session_state.do_cancel = True

        if st.session_state.get("do_confirm"):
            st.session_state.do_confirm = False
            f = st.session_state.form
            if f["name_mode"] == "existing" and f.get("_library_dim_id"):
                library = _fetch_dimension_library()
                entry = next((d for d in library if d["id"] == f["_library_dim_id"]), None)
                if entry is not None:
                    st.session_state.dims.append(_dim_from_library_entry(entry))
            elif f["name_mode"] == "custom" and f["name"].strip():
                new_dim = build_dim_from_form(f)
                new_dim["category"] = f["category"]
                st.session_state.dims.append(new_dim)
            st.session_state.form_open = False
            st.session_state.form = _empty_form()
            st.session_state.last_result = None
            st.rerun()

        if st.session_state.get("do_cancel"):
            st.session_state.do_cancel = False
            st.session_state.form_open = False
            st.session_state.form = _empty_form()
            st.rerun()

    st.markdown(
        '<div style="background:#1e1e30;border-radius:6px;padding:0.55rem 0.8rem;'
        'margin-top:0.75rem;font-size:0.82rem;color:#888">'
        'ℹ️ &nbsp; Set exactly one dimension as <b>Row</b> (mandatory). '
        'Set zero or one as <b>Column</b> (optional). '
        'All others become sheet-level dimensions.</div>',
        unsafe_allow_html=True,
    )


# ── Section: axis summary ─────────────────────────────────────────────────────

def render_axis_summary_section():
    st.markdown("### 📊 Table axes")
    st.caption(
        "These define the 2D grid (or 1D list, if no Column is set) inside each sheet. "
        "Change them any time using the Row/Column buttons above."
    )

    row_dim = _row_dim()
    col_dim = _col_dim()

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            st.markdown("**Row axis**")
            if row_dim:
                st.markdown(f"**{row_dim['name']}** `{row_dim['type']}`")
                st.caption(dim_summary(row_dim))
            else:
                st.caption("Not set — mark a dimension as Row above.")

    with c2:
        with st.container(border=True):
            st.markdown("**Column axis**")
            if col_dim:
                st.markdown(f"**{col_dim['name']}** `{col_dim['type']}`")
                st.caption(dim_summary(col_dim))
            elif row_dim:
                st.caption("Not set — sheets will be 1D (Row values only).")
            else:
                st.caption("Set a Row axis first.")


# ── Section: formula variables table ─────────────────────────────────────────

def _fv_slug_from_name(name: str) -> str:
    return get_value_slug(name.strip())


def _get_dim_slug_options() -> list[str]:
    slugs = []
    seen = set()
    row_dim = _row_dim()
    col_dim = _col_dim()

    for dim in st.session_state.dims:
        if dim.get("role") in ("row", "col"):
            continue
        if dim["name"] in _EXCLUDED_FROM_FORMULA:
            continue
        slug = get_dimension_slug(dim["name"])
        base = slug
        counter = 2
        while slug in seen:
            slug = f"{base}{counter}"
            counter += 1
        seen.add(slug)
        slugs.append(slug)

    if row_dim and row_dim["name"] not in _EXCLUDED_FROM_FORMULA:
        slug_name = get_dimension_slug(row_dim["name"])
        if slug_name not in seen:
            slugs.append(slug_name)
            seen.add(slug_name)
    if col_dim and col_dim["name"] not in _EXCLUDED_FROM_FORMULA:
        slug_name = get_dimension_slug(col_dim["name"])
        if slug_name not in seen:
            slugs.append(slug_name)
            seen.add(slug_name)

    for name in _ALL_DIMS:
        if name in _EXCLUDED_FROM_FORMULA:
            continue
        slug = _PREDEFINED_NAME_TO_SLUG.get(name, get_dimension_slug(name))
        if slug not in seen:
            seen.add(slug)
            slugs.append(slug)

    return slugs


def render_formula_variables_table():
    fvs: list = st.session_state.formula_variables

    hdr_col, add_col = st.columns([3, 2])
    with hdr_col:
        st.markdown("#### Formula Variables (Extra Inputs)")
        st.caption("Additional inputs used only in formulas (not part of the rate table).")
    with add_col:
        st.write("")
        if not st.session_state._fv_add_open and st.session_state._fv_editing is None:
            if st.button("＋ Add Variable", key="fv_open_add", use_container_width=True):
                st.session_state._fv_add_open = True
                st.session_state["_fv_new_name"] = ""
                st.session_state["_fv_new_default"] = ""
                st.rerun()

    if fvs:
        lh1, lh2, lh3, lh4 = st.columns([2.5, 2, 1.5, 1])
        with lh1:
            st.caption("Name")
        with lh2:
            st.caption("Slug")
        with lh3:
            st.caption("Default Value")
        with lh4:
            st.caption("Actions")

    to_delete = None
    for i, fv in enumerate(fvs):
        if st.session_state._fv_editing == i:
            with st.container(border=True):
                st.markdown(f"**Edit variable #{i + 1}**")
                ec1, ec2 = st.columns([5, 3])
                with ec1:
                    edit_name = st.text_input(
                        "Display Name", value=fv["name"],
                        key=f"fv_edit_name_{i}",
                    )
                    edit_slug = _fv_slug_from_name(edit_name) if edit_name else fv["slug"]
                    st.caption(f"Slug: `{edit_slug}`")
                with ec2:
                    edit_default = st.text_input(
                        "Default Value", value=str(fv.get("default_value", "")),
                        key=f"fv_edit_default_{i}",
                        placeholder="—",
                    )

                existing_slugs = set(_get_dim_slug_options())
                other_fv_slugs = {fvs[j]["slug"] for j in range(len(fvs)) if j != i}
                if edit_slug and (edit_slug in existing_slugs or edit_slug in other_fv_slugs):
                    st.warning(f"⚠️ Slug `{edit_slug}` conflicts with an existing variable.")

                sb1, sb2, _ = st.columns([1, 1, 4])
                with sb1:
                    if st.button("Save", key=f"fv_save_{i}", type="primary"):
                        if edit_name.strip() and edit_slug.strip():
                            fvs[i] = {
                                "name": edit_name.strip(),
                                "slug": edit_slug.strip(),
                                "default_value": edit_default.strip(),
                            }
                            st.session_state._fv_editing = None
                            st.rerun()
                with sb2:
                    if st.button("Cancel", key=f"fv_cancel_edit_{i}"):
                        st.session_state._fv_editing = None
                        st.rerun()
        else:
            with st.container(border=True):
                dc1, dc2, dc3, dc4 = st.columns([2.5, 2, 1.5, 1])
                with dc1:
                    st.markdown(f"**{fv['name']}**")
                with dc2:
                    st.markdown(
                        f'<span class="fv-slug">{fv["slug"]}</span>',
                        unsafe_allow_html=True,
                    )
                with dc3:
                    dv = fv.get("default_value", "")
                    st.markdown(
                        f'<small style="color:#888">{dv if dv != "" else "—"}</small>',
                        unsafe_allow_html=True,
                    )
                with dc4:
                    btn_c1, btn_c2 = st.columns(2)
                    with btn_c1:
                        if st.button("✏️", key=f"fv_edit_btn_{i}", help="Edit"):
                            st.session_state._fv_editing = i
                            st.session_state._fv_add_open = False
                            st.rerun()
                    with btn_c2:
                        if st.button("🗑", key=f"fv_del_btn_{i}", help="Delete"):
                            to_delete = i

    if to_delete is not None:
        fvs.pop(to_delete)
        if st.session_state._fv_editing == to_delete:
            st.session_state._fv_editing = None
        st.rerun()

    if st.session_state._fv_add_open:
        with st.container(border=True):
            st.markdown("**New formula variable**")
            ac1, ac2 = st.columns([5, 3])
            with ac1:
                new_name = st.text_input(
                    "Display Name", value="",
                    key="fv_add_name",
                    placeholder="e.g. Loan Amount",
                )
                new_slug = _fv_slug_from_name(new_name) if new_name else ""
                st.caption(f"Slug: `{new_slug}`")
            with ac2:
                new_default = st.text_input(
                    "Default Value", value="",
                    key="fv_add_default",
                    placeholder="—",
                )

            existing_slugs = set(_get_dim_slug_options())
            existing_fv_slugs = {fv["slug"] for fv in fvs}
            if new_slug and (new_slug in existing_slugs or new_slug in existing_fv_slugs):
                st.warning(f"⚠️ Slug `{new_slug}` conflicts with an existing variable.")

            ab1, ab2, _ = st.columns([1, 1, 4])
            with ab1:
                if st.button("Add variable", key="fv_confirm_add", type="primary"):
                    if new_name.strip() and new_slug.strip():
                        fvs.append({
                            "name": new_name.strip(),
                            "slug": new_slug.strip(),
                            "default_value": new_default.strip(),
                        })
                        st.session_state._fv_add_open = False
                        st.rerun()
            with ab2:
                if st.button("Cancel", key="fv_cancel_add"):
                    st.session_state._fv_add_open = False
                    st.rerun()

    if not fvs and not st.session_state._fv_add_open:
        st.caption("No formula variables yet. Add one above.")


# ── Section: formula builder ──────────────────────────────────────────────────

def render_formula_section():
    st.markdown("### 🧮 Calculation logic")
    st.caption(
        "Build the premium formula by editing the box directly or inserting tokens. "
        "`rater_val` is always the value fetched from the rate table."
    )

    if "formula_editor_input" not in st.session_state:
        st.session_state.formula_editor_input = " ".join(st.session_state.get("formula_tokens", []))

    def backspace_callback():
        val = st.session_state.formula_editor_input.strip()
        if " " in val:
            parts = val.split()
            parts.pop()
            st.session_state.formula_editor_input = " ".join(parts)
        elif val:
            st.session_state.formula_editor_input = val[:-1]
        else:
            st.session_state.formula_editor_input = ""

    def clear_callback():
        st.session_state.formula_editor_input = ""

    def insert_var_callback():
        chosen_var = st.session_state.formula_var_select
        current = st.session_state.formula_editor_input.strip()
        if current:
            st.session_state.formula_editor_input = f"{current} {chosen_var}"
        else:
            st.session_state.formula_editor_input = chosen_var

    def insert_op_callback(op: str):
        current = st.session_state.formula_editor_input.strip()
        if current:
            st.session_state.formula_editor_input = f"{current} {op}"
        else:
            st.session_state.formula_editor_input = op

    bar_col, bs_col, clr_col = st.columns([11, 1, 1])

    with bar_col:
        st.text_input(
            label="Formula Editor",
            label_visibility="collapsed",
            placeholder="formula will appear here… or type directly",
            key="formula_editor_input",
        )

    with bs_col:
        st.button(
            "⌫",
            key="formula_backspace",
            help="Remove last part/character",
            use_container_width=True,
            on_click=backspace_callback,
        )

    with clr_col:
        st.button(
            "✕",
            key="formula_clear",
            help="Clear all",
            use_container_width=True,
            on_click=clear_callback,
        )

    st.write("")

    palette_col1, palette_col2 = st.columns(2)

    with palette_col1:
        st.markdown("**Variables**")
        fv_slugs_list = [fv["slug"] for fv in st.session_state.get("formula_variables", [])]
        var_options = ["rater_val"] + _get_dim_slug_options() + fv_slugs_list

        st.selectbox(
            "variable", var_options,
            label_visibility="collapsed",
            key="formula_var_select",
        )
        st.button(
            "Insert variable",
            key="insert_var",
            on_click=insert_var_callback,
        )

    with palette_col2:
        st.markdown("**Operators**")
        op_cols = st.columns(6)
        for idx, op in enumerate(["+", "-", "*", "/", "(", ")"]):
            with op_cols[idx]:
                st.button(
                    op,
                    key=f"op_btn_{op}",
                    on_click=insert_op_callback,
                    args=(op,),
                )

    st.write("")

    set_col, status_col = st.columns([2, 2])
    
    with set_col:
        if st.button("Set formula", type="primary", key="set_formula_btn"):
            formula_str = st.session_state.formula_editor_input
            if not formula_str.strip():
                st.session_state["_formula_status"] = ("empty", "No formula entered")
            else:
                dim_slugs = set(_get_dim_slug_options())
                fv_slugs = {fv["slug"] for fv in st.session_state.get("formula_variables", [])}
                allowed = dim_slugs | fv_slugs | {"rater_val"}
                try:
                    validate_formula(formula_str, allowed)
                    st.session_state["_formula_status"] = ("success", "Formula is valid")
                except FormulaError as e:
                    st.session_state["_formula_status"] = ("error", str(e))
            st.rerun()

    with status_col:
        status = st.session_state.get("_formula_status")
        if status:
            status_type, status_msg = status
            if status_type == "success":
                st.success(f"✅ {status_msg}")
            elif status_type == "error":
                st.error(f"❌ {status_msg}")
            else:
                st.warning(f"⚠️ {status_msg}")      


# ── Build result JSON ─────────────────────────────────────────────────────────

def _parse_dim_to_dict(dim_type: str, dim_config: dict) -> dict:
    op_map = {
        "<":  "Less than",
        "<=": "Less than or equal to",
        ">":  "Greater than",
        ">=": "Greater than or equal to",
        "=":  "Equal to",
    }
    result = {}

    if dim_type == "Enum":
        for val in dim_config.get("values", []):
            slug = get_value_slug(str(val))
            result[slug] = str(val)

    elif dim_type == "Range":
        lo = dim_config.get("min", 0)
        hi = dim_config.get("max", 100)
        result[f"{lo}_{hi}"] = f"{lo} to {hi}"

    elif dim_type == "Comparison":
        for comp in dim_config.get("comparisons", []):
            op  = comp.get("op", "")
            val = comp.get("val", "")
            op_slug  = operator_to_slug(op)
            val_slug = get_value_slug(val)
            slug  = f"{op_slug}_{val_slug}".strip("_")
            label = f"{op_map.get(op, op)} {val}".strip()
            result[slug] = label

    return result


def _axis_def(dim):
    if dim is None:
        return None
    if dim["type"] == "Range":
        return {
            "name": dim["name"], "type": "Range",
            "min": dim["config"]["min"], "max": dim["config"]["max"],
        }
    if dim["type"] == "Enum":
        return {"name": dim["name"], "type": "Enum", "config": dim["config"].get("values", [])}
    if dim["type"] == "Comparison":
        return {"name": dim["name"], "type": "Comparison", "config": dim["config"].get("comparisons", [])}
    return None


def _build_result() -> dict:
    dimensions_def = {}
    parsed_dims = {}

    outer_dims = [d for d in st.session_state.dims if d.get("role") not in ("row", "col")]
    row_dim = _row_dim()
    col_dim = _col_dim()

    for dim in outer_dims:
        dim_name   = dim["name"]
        dim_type   = dim["type"]
        dim_config = dim["config"]

        dim_slug = get_dimension_slug(dim_name)
        base_slug = dim_slug
        counter = 2
        while dim_slug in parsed_dims:
            dim_slug = f"{base_slug}{counter}"
            counter += 1

        if dim_type == "Enum":
            dimensions_def[dim_name] = {"type": "Enum", "config": dim_config.get("values", [])}
        elif dim_type == "Range":
            dimensions_def[dim_name] = {
                "type": "Range",
                "config": {"min": dim_config.get("min"), "max": dim_config.get("max")},
            }
        elif dim_type == "Comparison":
            dimensions_def[dim_name] = {
                "type": "Comparison",
                "config": dim_config.get("comparisons", []),
            }

        parsed_dims[dim_slug] = _parse_dim_to_dict(dim_type, dim_config)

    row_axis_def = _axis_def(row_dim)
    if row_axis_def is not None:
        row_axis_def["enabled"] = True
    col_axis_def = _axis_def(col_dim)

    formula_str = st.session_state.get("formula_editor_input", "")

    return {
        "definition": {
            "dimensions": dimensions_def,
            "axes": {
                "row": row_axis_def,
                "col": col_axis_def,
            },
        },
        "parsed_dimensions": parsed_dims,
        "calculation": {
            "formula": formula_str,
            "constants": [],
            "formula_variables": [dict(fv) for fv in st.session_state.get("formula_variables", [])],
        },
        "rounding_rule": st.session_state.get("rounding_rule", "nearest"),
    }


# ── Section: generate ─────────────────────────────────────────────────────────

def render_generate_section():
    st.divider()

    row_dim = _row_dim()
    col_dim = _col_dim()
    outer_count = len([d for d in st.session_state.dims if d.get("role") not in ("row", "col")])

    if row_dim:
        axis_caption = f"{row_dim['name']} (row)"
    else:
        axis_caption = "no row axis (missing)"
    axis_caption += f" · {col_dim['name']} (column)" if col_dim else " · no column (1D)"
    st.caption(f"{outer_count} sheet dimension(s) · {axis_caption}")

    if st.button("Generate structure & save template", type="primary"):
        has_errors = False
        if not st.session_state.template_name.strip():
            st.error("IC/product name is not provided. Enter a name before generating.", icon="⚠️")
            has_errors = True
        if row_dim is None:
            st.error("Row dimension is not set. Mark a dimension as Row before generating.", icon="⚠️")
            has_errors = True

        if not has_errors:
            result = _build_result()
            st.session_state.last_result = result

            payload = {**result, "name": st.session_state.template_name.strip()}
            tid = st.session_state.last_template_id

            try:
                if tid:
                    resp = requests.put(
                        f"{SERVER_URL}/templates/{tid}",
                        json=payload,
                        timeout=10,
                    )
                else:
                    resp = requests.post(
                        f"{SERVER_URL}/templates",
                        json=payload,
                        timeout=10,
                    )
                if resp.ok:
                    data = resp.json()
                    st.session_state.last_template_id = data["template_id"]
                    st.success(f"Template saved! **{data['name']}** · ID: `{data['template_id']}`")
                else:
                    st.error(f"Server error: {resp.text}")
            except requests.exceptions.ConnectionError:
                st.warning(
                    "Could not reach the Flask server. "
                    "Start it with `python server.py` and try again.",
                    icon="⚠️",
                )

            st.rerun()

    if st.session_state.last_result:
        result = st.session_state.last_result

        st.success("Structure generated!")

        with st.expander("View JSON"):
            st.json(result)

        st.divider()

        with st.spinner("Building Excel workbook…"):
            try:
                excel_data = {**result, "template_id": st.session_state.last_template_id}
                xlsx_bytes = generate_excel(excel_data)

                n_sheets = max(1, 1)
                for vals in result["parsed_dimensions"].values():
                    n_sheets *= len(vals)

                row_axis = result["definition"]["axes"]["row"]
                col_axis = result["definition"]["axes"]["col"]
                n_rows = row_axis["max"] - row_axis["min"] + 1 if row_axis.get("type") == "Range" else None

                if col_axis:
                    n_cols = col_axis["max"] - col_axis["min"] + 1 if col_axis.get("type") == "Range" else None
                    grid_caption = f"**{n_rows or '?'}** rows × **{n_cols or '?'}** columns each"
                else:
                    grid_caption = f"**{n_rows or '?'}**-row 1D list each ({row_axis['name']} only)"

                st.info(f"📊 **{n_sheets}** data sheet(s) · {grid_caption}", icon="📋")

                st.download_button(
                    label="⬇️ Download Excel (.xlsx)",
                    data=xlsx_bytes,
                    file_name=f"{st.session_state.template_name}_rater_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
            except Exception as e:
                st.error(f"Excel generation failed: {e}")


# ── Tab 0: Dimension Library (internal admin) ────────────────────────────────
# Master repository of reusable dimensions, consumed by Tab 1's "Choose
# Existing Dimension" flow. System dimensions (is_system) are immutable in
# name/type — no rename/retype UI is offered for them, only their
# configuration can change. Only admin-created dimensions can be edited
# (fully) or deleted.

def _library_row_summary(dim: dict) -> str:
    """Table-row display for a library dimension's configuration. Enum
    values are truncated to the first 3 with a '(+N more)' suffix once there
    are more than 3 — the full list still loads normally into the edit form.
    Range/Comparison reuse the same summary as the Template Builder table."""
    if dim["type"] == "Enum":
        values = dim["config"].get("values", [])
        if not values:
            return "—"
        if len(values) <= 3:
            return ", ".join(values)
        return f"{', '.join(values[:3])}... (+{len(values) - 3} more)"
    return dim_summary({"type": dim["type"], "config": dim["config"]})


def _open_library_add_form():
    st.session_state.lib_form_open = True
    st.session_state.lib_editing_id = None
    st.session_state.lib_editing_is_system = False
    st.session_state.lib_form = _empty_library_form()
    st.session_state.lib_delete_confirm_id = None
    st.session_state.lib_form_nonce += 1


def _open_library_editor(dim: dict):
    f = _empty_library_form()
    f["name"] = dim["name"]
    f["type"] = dim["type"]
    if dim["type"] == "Enum":
        f["values"] = list(dim["config"].get("values", []))
    elif dim["type"] == "Range":
        f["min"] = dim["config"].get("min", 0)
        f["max"] = dim["config"].get("max", 100)
    elif dim["type"] == "Comparison":
        f["comparisons"] = list(dim["config"].get("comparisons", []))

    st.session_state.lib_form_open = True
    st.session_state.lib_editing_id = dim["id"]
    st.session_state.lib_editing_is_system = dim["is_system"]
    st.session_state.lib_form = f
    st.session_state.lib_delete_confirm_id = None
    st.session_state.lib_form_nonce += 1


def _close_library_form():
    st.session_state.lib_form_open = False
    st.session_state.lib_editing_id = None
    st.session_state.lib_editing_is_system = False
    st.session_state.lib_form = _empty_library_form()


def _render_library_form_body(key_ns: str):
    """Renders the Add/Edit form body. key_ns namespaces every widget key
    inside it so reopening a form (even the same row) never inherits a
    stale, unsaved value from a previous open."""
    f = st.session_state.lib_form
    editing_id = st.session_state.lib_editing_id
    editing_is_system = st.session_state.lib_editing_is_system

    with st.container(border=True):
        if editing_id:
            title = "Edit System Dimension" if editing_is_system else "Edit Custom Dimension"
        else:
            title = "Add Custom Dimension"
        st.markdown(f"##### {title}")
        if editing_is_system:
            st.caption("Name and type are locked for system dimensions — only the configuration can change.")

        col_name, col_type = st.columns(2)
        with col_name:
            f["name"] = st.text_input(
                "Name", value=f["name"], key=f"{key_ns}_name",
                placeholder="e.g. Occupation", disabled=editing_is_system,
            )
        with col_type:
            f["type"] = st.selectbox(
                "Type", ["Enum", "Range", "Comparison"],
                index=["Enum", "Range", "Comparison"].index(f["type"]),
                key=f"{key_ns}_type", disabled=editing_is_system,
            )

        if f["type"] == "Enum":
            _render_enum_inputs(key_ns, f)
        elif f["type"] == "Range":
            c1, c2 = st.columns(2)
            with c1:
                f["min"] = st.number_input("Min", value=int(f["min"]), step=1, key=f"{key_ns}_min")
            with c2:
                f["max"] = st.number_input("Max", value=int(f["max"]), step=1, key=f"{key_ns}_max")
        elif f["type"] == "Comparison":
            _render_comparison_inputs(key_ns, f)

        st.write("")
        bc1, bc2, _ = st.columns([3, 2, 3])
        with bc1:
            if st.button("Save", key=f"{key_ns}_confirm", type="primary", use_container_width=True):
                name_clean = f["name"].strip()
                if not name_clean:
                    st.error("Name is required.", icon="⚠️")
                else:
                    if f["type"] == "Enum":
                        config = {"values": list(f["values"])}
                    elif f["type"] == "Range":
                        config = {"min": f["min"], "max": f["max"]}
                    else:
                        config = {"comparisons": [dict(c) for c in f["comparisons"]]}

                    if editing_id:
                        res = _update_library_dimension(editing_id, name_clean, f["type"], config)
                    else:
                        res = _create_library_dimension(name_clean, f["type"], config)

                    if res["ok"]:
                        _close_library_form()
                        st.session_state.lib_message = ("success", f"Saved '{name_clean}'.")
                        st.rerun()
                    else:
                        st.error(res["error"], icon="⚠️")
        with bc2:
            if st.button("Cancel", key=f"{key_ns}_cancel", use_container_width=True):
                _close_library_form()
                st.rerun()


def render_tab_dimension_library():
    st.markdown("### Dimension Library")
    st.caption(
        "The master list of reusable dimensions available to every product template. "
        "System dimensions are built-in defaults — their name and type are locked, "
        "but their configuration can still be updated. Admin-created dimensions can be "
        "fully added, edited, and deleted."
    )

    msg = st.session_state.lib_message
    if msg:
        level, text = msg
        (st.success if level == "success" else st.error)(text, icon="✅" if level == "success" else "⚠️")
        st.session_state.lib_message = None

    library = _fetch_dimension_library()

    hdr_col, btn_col = st.columns([3, 2])
    with hdr_col:
        st.markdown("#### Dimensions")
    with btn_col:
        st.write("")
        if st.button("＋ Add Library Dimension", key="lib_open_form", use_container_width=True):
            _open_library_add_form()
            st.rerun()

    if st.session_state.lib_form_open and st.session_state.lib_editing_id is None:
        _render_library_form_body(key_ns=f"libform_new_{st.session_state.lib_form_nonce}")

    st.text_input(
        "Search dimensions", key="lib_search",
        placeholder="🔍 Search dimensions...", label_visibility="collapsed",
    )

    if not library:
        st.caption("No dimensions found (or the server is unreachable).")
    else:
        query = st.session_state.lib_search.strip().lower()
        visible = [d for d in library if query in d["name"].lower()] if query else library

        if query and not visible:
            st.caption("No dimensions found.")
        else:
            lh1, lh2, lh3, lh4, lh5 = st.columns([2.5, 3, 1.3, 1.3, 1.9])
            for col, label in zip((lh1, lh2, lh3, lh4, lh5), ["Name", "Values", "Type", "Status", ""]):
                with col:
                    st.caption(label)

            for dim in visible:
                with st.container(border=True):
                    c1, c2, c3, c4, c5 = st.columns([2.5, 3, 1.3, 1.3, 1.9])
                    with c1:
                        st.markdown(f"**{dim['name']}**")
                    with c2:
                        st.markdown(f"<small>{_library_row_summary(dim)}</small>", unsafe_allow_html=True)
                    with c3:
                        st.markdown(f"<small>{dim['type']}</small>", unsafe_allow_html=True)
                    with c4:
                        if dim["is_system"]:
                            st.markdown('<span class="pill pill-blue">🔵 System</span>', unsafe_allow_html=True)
                        else:
                            st.markdown('<span class="pill pill-green">🟢 Custom</span>', unsafe_allow_html=True)
                    with c5:
                        if dim["is_system"]:
                            if st.button("✏️ Edit", key=f"lib_edit_{dim['id']}", help="Edit configuration",
                                         use_container_width=True):
                                _open_library_editor(dim)
                                st.rerun()
                        else:
                            ec1, ec2 = st.columns(2)
                            with ec1:
                                if st.button("✏️", key=f"lib_edit_{dim['id']}", help="Edit"):
                                    _open_library_editor(dim)
                                    st.rerun()
                            with ec2:
                                if st.button("🗑", key=f"lib_del_{dim['id']}", help="Delete"):
                                    st.session_state.lib_delete_confirm_id = dim["id"]
                                    st.rerun()

                    if st.session_state.lib_delete_confirm_id == dim["id"]:
                        st.warning(f"Delete '{dim['name']}' from the library?", icon="⚠️")
                        dc1, dc2, dc3 = st.columns([1, 1, 3])
                        with dc1:
                            if st.button("Confirm delete", key=f"lib_del_confirm_{dim['id']}", type="primary"):
                                res = _delete_library_dimension(dim["id"])
                                if not res["ok"] and res.get("usage_count"):
                                    st.session_state[f"_lib_force_{dim['id']}"] = True
                                    st.error(res["error"], icon="⚠️")
                                elif not res["ok"]:
                                    st.error(res["error"], icon="⚠️")
                                else:
                                    st.session_state.lib_delete_confirm_id = None
                                    st.session_state.lib_message = ("success", f"Deleted '{dim['name']}'.")
                                    st.rerun()
                        with dc2:
                            if st.button("Cancel", key=f"lib_del_cancel_{dim['id']}"):
                                st.session_state.lib_delete_confirm_id = None
                                st.rerun()
                        if st.session_state.get(f"_lib_force_{dim['id']}"):
                            st.caption("This dimension appears to be used by a saved template.")
                            if st.button("Delete anyway", key=f"lib_del_force_{dim['id']}"):
                                res = _delete_library_dimension(dim["id"], force=True)
                                if res["ok"]:
                                    st.session_state.lib_delete_confirm_id = None
                                    st.session_state[f"_lib_force_{dim['id']}"] = False
                                    st.session_state.lib_message = ("success", f"Deleted '{dim['name']}'.")
                                    st.rerun()
                                else:
                                    st.error(res["error"], icon="⚠️")

                if st.session_state.lib_form_open and st.session_state.lib_editing_id == dim["id"]:
                    _render_library_form_body(key_ns=f"libform_edit_{dim['id']}_{st.session_state.lib_form_nonce}")


# ── Tab 1: Template Builder ───────────────────────────────────────────────────

def render_tab_template_builder():
    if st.session_state.last_template_id and not st.session_state.template_name:
        detail = _fetch_template_detail(st.session_state.last_template_id)
        if detail:
            st.session_state.template_name = detail.get("name", "")
            st.session_state.rounding_rule = detail.get("rounding_rule", "nearest")
    
    p_name, p_code = st.columns(2)

    with p_name:
        st.text_input(
            'Product name',
            key="template_name",
            placeholder="e.g. Motor Comprehensive",
        )

    with p_code:
        st.text_input(
            'Product Code',
            key="product_code",
            placeholder="e.g. ICICI001",
        )

    rounding_options = {
        "nearest": "Round to nearest whole rupee (default)",
        "up": "Round up (ceiling)",
        "down": "Round down (floor)",
        "none": "No rounding — compare exact values",
    }
    with st.container(border=True):
        st.markdown("🎯 **Premium Rounding (for comparison)**")
        st.caption(
            "Only affects the MATCH/MISMATCH verdict shown in Premium Comparison — "
            "never changes the calculated premium itself. Most partner APIs return "
            "whole-rupee premiums while the local calculation keeps decimals, so "
            "'Round to nearest' is the right default for most templates."
        )
        st.selectbox(
            "Rounding rule",
            options=list(rounding_options.keys()),
            format_func=lambda k: rounding_options[k],
            key="rounding_rule",
            label_visibility="collapsed",
        )

    render_dims_section()
    st.divider()
    render_axis_summary_section()
    st.divider()
    render_formula_variables_table()
    st.divider()
    render_formula_section()
    render_generate_section()


# ── Tab 2: Upload & Flatten ───────────────────────────────────────────────────

def render_tab_upload():
    st.markdown("### Upload & Flatten")
    st.caption(
        "Upload a filled-in Excel file to store its rate values in the database. "
        "The template is identified automatically from the workbook's hidden _meta sheet."
    )

    uploaded_file = st.file_uploader(
        "Upload filled Excel (.xlsx)", type=["xlsx"], key="upload_file"
    )

    if st.button("Upload & Flatten", type="primary", key="do_flatten"):
        if not uploaded_file:
            st.error("Upload an Excel file first.")
        else:
            with st.spinner("Flattening…"):
                try:
                    resp = requests.post(
                        f"{SERVER_URL}/flatten",
                        files={"file": (uploaded_file.name, uploaded_file.getvalue(),
                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                        timeout=120,
                    )
                    if resp.ok:
                        data = resp.json()
                        st.success(
                            f"✅ **{data['cells_written']}** cells written · "
                            f"**{data['blank_cells_count']}** blank · "
                            f"**{data['completion_pct']}%** complete"
                        )
                        if data.get("blank_cells"):
                            with st.expander(f"Blank cells ({len(data['blank_cells'])} shown)"):
                                st.json(data["blank_cells"])
                    else:
                        st.error(f"Server error: {resp.text}")
                except requests.exceptions.ConnectionError:
                    st.warning("Could not reach the Flask server.", icon="⚠️")


# ── Tab 3: Calculate Premium ──────────────────────────────────────────────────

def _fetch_templates() -> list[dict]:
    try:
        resp = requests.get(f"{SERVER_URL}/templates", timeout=10)
        if resp.ok:
            return resp.json()
    except requests.exceptions.ConnectionError:
        pass
    return []


def _fetch_template_detail(template_id: str) -> dict | None:
    try:
        resp = requests.get(f"{SERVER_URL}/templates/{template_id}", timeout=10)
        if resp.ok:
            return resp.json()
    except requests.exceptions.ConnectionError:
        pass
    return None


def _render_calc_inputs(full_def: dict) -> dict:
    definition = full_def.get("definition", {})
    dimensions = definition.get("dimensions", {})
    axes = definition.get("axes", {})
    row_axis = axes.get("row")
    col_axis = axes.get("col")
    inputs = {}

    from datetime import date

    for dim_name, dim_def in dimensions.items():
        dim_slug = get_dimension_slug(dim_name)
        dim_type = dim_def["type"]

        if dim_slug == "age":
            dob_val = st.date_input(
                "Date of Birth (DOB)",
                value=date(1999, 1, 1),
                key=f"calc_input_{dim_slug}_dob"
            )
            inputs["dob"] = dob_val.strftime("%d-%m-%Y")
            continue

        if dim_type == "Enum":
            options = dim_def.get("config", [])
            val = st.selectbox(dim_name, options, key=f"calc_input_{dim_slug}")
        elif dim_type == "Range":
            lo = dim_def["config"]["min"]
            hi = dim_def["config"]["max"]
            val = st.number_input(dim_name, min_value=int(lo), max_value=int(hi),
                                  value=int(lo), step=1, key=f"calc_input_{dim_slug}")
        elif dim_type == "Comparison":
            val = st.number_input(dim_name, value=0.0, step=1.0, key=f"calc_input_{dim_slug}")
        else:
            val = st.text_input(dim_name, key=f"calc_input_{dim_slug}")

        inputs[dim_slug] = val

    if row_axis:
        row_slug = get_dimension_slug(row_axis["name"])
        if row_slug == "age":
            dob_val = st.date_input(
                "DOB (row)",
                value=date(1999, 1, 1),
                key="calc_row_val_dob",
            )
            inputs["dob"] = dob_val.strftime("%d-%m-%Y")
        else:
            rtype = row_axis.get("type", "Range")
            if rtype == "Range":
                row_val = st.number_input(
                    row_axis["name"] + " (row)",
                    min_value=int(row_axis["min"]), max_value=int(row_axis["max"]),
                    value=int(row_axis["min"]), step=1, key="calc_row_val",
                )
            elif rtype == "Enum":
                row_val = st.selectbox(row_axis["name"] + " (row)", row_axis.get("config", []), key="calc_row_val")
            else:
                row_val = st.text_input(row_axis["name"] + " (row)", key="calc_row_val")
            inputs[row_slug] = row_val

    if col_axis:
        col_slug = get_dimension_slug(col_axis["name"])
        if col_slug == "age":
            dob_val = st.date_input(
                "DOB (col)",
                value=date(1999, 1, 1),
                key="calc_col_val_dob",
            )
            inputs["dob"] = dob_val.strftime("%d-%m-%Y")
        else:
            ctype = col_axis.get("type", "Range")
            if ctype == "Range":
                col_val = st.number_input(
                    col_axis["name"] + " (col)",
                    min_value=int(col_axis["min"]), max_value=int(col_axis["max"]),
                    value=int(col_axis["min"]), step=1, key="calc_col_val",
                )
            elif ctype == "Enum":
                col_val = st.selectbox(col_axis["name"] + " (col)", col_axis.get("config", []), key="calc_col_val")
            else:
                col_val = st.text_input(col_axis["name"] + " (col)", key="calc_col_val")
            inputs[col_slug] = col_val

    calculation = full_def.get("calculation", {})
    for fv in calculation.get("formula_variables", []):
        fv_slug = fv["slug"]
        fv_default = fv.get("default_value", "")
        try:
            default_num = int(fv_default) if fv_default not in (None, "") else 0.0
        except ValueError:
            default_num = 1
        val = st.number_input(
            fv["name"],
            value=default_num,
            step=1,
            key=f"calc_input_fv_{fv_slug}",
        )
        inputs[fv_slug] = val
    
    formula_str = full_def.get("calculation", {}).get("formula", "") or ""
    for slug, display_name in _PREDEFINED_SLUG_TO_NAME.items():
        if slug in inputs:
            continue
        if not re.search(rf"\b{re.escape(slug)}\b", formula_str):
            continue
        val = st.number_input(
            display_name,
            value=0,
            step=1,
            key=f"calc_input_predef_{slug}",
        )
        inputs[slug] = val

    return inputs


# ── Flow execution calls ─────────────────────────────────────────────────────

# ── Custom payload (per flow_type) ───────────────────────────────────────────

def _fetch_payload_template(flow_type: str) -> dict | None:
    """Fetch the server's default hardcoded payload template for flow_type,
    used to pre-fill the 'Edit Payload' popup. Returns None on failure."""
    try:
        resp = requests.get(f"{SERVER_URL}/api/payload-template/{flow_type}", timeout=10)
        if resp.ok:
            return resp.json()
    except requests.exceptions.RequestException:
        pass
    return None


def _extract_lead_fields_from_payload(payload: dict) -> dict:
    """Pull name/mobile/email out of a (custom) payload's proposer or
    first borrower, so the 'Lead Creation Parameters' UI fields can be
    kept in sync with whatever a custom payload actually contains."""
    if not isinstance(payload, dict):
        return {}

    source = None
    if isinstance(payload.get("proposer"), dict):
        source = payload["proposer"]
    elif isinstance(payload.get("borrowers"), list) and payload["borrowers"]:
        source = payload["borrowers"][0]

    if not source:
        return {}

    first = (source.get("first_name") or "").strip()
    last = (source.get("last_name") or "").strip()
    name = (first + " " + last).strip()

    return {
        "name": name,
        "mobile": (source.get("phone_number") or "").strip(),
        "email": (source.get("email") or "").strip(),
    }


@st.dialog("Choose Workflow",width="large")
def _workflow_setup_dialog(mode_key: str):
    st.caption(
        "Pick which partner API this run should use, then optionally edit "
        "its payload before continuing."
    )
    flow_type_label=st.radio("Choose API",["Save Loan","Save Lead","Create Lead"],key=f"_workflow_setup_radio_{mode_key}")
    flow_type_map={"Save Loan":"save_loan","Save Lead":"save_lead","Create Lead":"create_lead"}
    flow_type=flow_type_map[flow_type_label]
    st.divider()

    existing = st.session_state.custom_payloads.get(flow_type)
    if existing is not None:
        st.info("A custom payload is currently active for this workflow.", icon="✏️")
        starting_text = json.dumps(existing, indent=2)
    else:
        default_template = _fetch_payload_template(flow_type)
        if default_template is None:
            st.warning(
                "Could not fetch the default template from the server "
                "(is it running?). You can still paste a payload manually.",
                icon="⚠️",
            )
            default_template = {}
        starting_text = json.dumps(default_template, indent=2)

    text_key = f"_workflow_setup_editor_{mode_key}_{flow_type}"
    edited_text = st.text_area(
        "Payload JSON", value=starting_text, height=350,
        key=text_key, label_visibility="collapsed",
    )
    c1,c2=st.columns(2)
    with c1:
        if st.button("Save & Continue", type="primary", use_container_width=True, key=f"_workflow_setup_save_{mode_key}"):
            try:
                parsed=json.loads(edited_text)
            except json.JSONDecodeError as e:
                st.error(f"Invalid JSON: {e}")
            else:
                if not isinstance(parsed, dict):
                    st.error("Payload must be a JSON object.")
                else:
                    st.session_state.custom_payloads[flow_type]=parsed
                    if flow_type in ("save_lead","create_lead"):
                        extracted=_extract_lead_fields_from_payload(parsed)
                        if extracted.get("name"):
                            st.session_state["calc_lead_name"]=extracted["name"]
                        if extracted.get("mobile"):
                            st.session_state["calc_lead_mobile"]=extracted["mobile"]
                        if extracted.get("email"):
                            st.session_state["calc_lead_email"]=extracted["email"]
                    st.session_state.calc_flow_type[mode_key] = flow_type
                    st.rerun()
    with c2:
        if st.button("Keep Default and Continue",use_container_width=True,key=f"_workflow_setup_default_{mode_key}"):
            st.session_state.calc_flow_type[mode_key] = flow_type
            st.rerun()

def _initialize_api_session(
    flow_type: str,
    lead_details: dict = None,
    inputs: dict | None = None,
    custom_payload: dict | None = None,
) -> dict:
    """Step 1: Init workflow session and fetch the session token or premium result."""
    endpoint_map = {
        "save_loan": "/api/get-premium",
        "save_lead": "/api/save-lead",
        "create_lead": "/api/create-lead",
    }
    endpoint = endpoint_map.get(flow_type, "/api/get-premium")
    
    payload = {
        "dimensions": inputs or {},
    }
    if flow_type == "save_loan":
        payload["flow_type"] = flow_type
    if flow_type in ["save_lead", "create_lead"] and lead_details:
        payload["lead_details"] = lead_details
    if custom_payload:
        # Base payload the user supplied instead of the hardcoded template.
        # Rater dimensions + lead details above are still applied on top of
        # it server-side.
        payload["payload"] = custom_payload

    try:
        resp = requests.post(
            f"{SERVER_URL}{endpoint}",
            json=payload,
            timeout=10,
        )
        if resp.ok:
            data = resp.json()
            token = data.get("session_token") or data.get("lead_id") or data.get("transaction_id") or data.get("loan_id")

            if not token:
                try:
                    err_data = resp.json()
                except ValueError:
                    err_data = {"error": resp.text}
                return {"ok": False, "error": err_data.get("error", "No session token returned.")}

            return {"ok": True, "session_token": token, "data": data}

        try:
            err_data = resp.json()
        except ValueError:
            err_data = {"error": resp.text}
        return {"ok": False, "error": err_data.get("error", resp.text)}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": f"Could not reach {endpoint} server."}
    except requests.exceptions.RequestException as e:
        return {"ok": False, "error": str(e)}


def _calculate_premium_with_token(tid: str, inputs: dict, session_token: str) -> dict:
    """Step 2: Unified rater evaluation passing session token alongside calculation variables."""
    try:
        payload = {
            "inputs": inputs,
            "session_token": session_token
        }
        resp = requests.post(
            f"{SERVER_URL}/templates/{tid}/calculate",
            json=payload,
            timeout=10,
        )
        if resp.ok:
            return {"ok": True, "data": resp.json()}
        try:
            err_data = resp.json()
        except ValueError:
            err_data = {"error": resp.text}
        return {"ok": False, "error": err_data.get("error", resp.text)}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": "Could not reach premium evaluation server."}


def _start_bulk_premium_job(
    tid: str,
    file_bytes: bytes,
    filename: str,
    flow_type: str,
    custom_payload: dict | None = None,
) -> dict:
    """Start a bulk processing job and return the job id."""
    try:
        form_data = {"flow_type": flow_type}
        if custom_payload:
            form_data["payload"] = json.dumps(custom_payload)
        resp = requests.post(
            f"{SERVER_URL}/templates/{tid}/calculate-bulk",
            files={"file": (filename, file_bytes,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data=form_data,
            timeout=30,
        )
        if resp.ok:
            data = resp.json()
            return {"ok": True, "job_id": data.get("job_id"), "status_url": data.get("status_url")}
        try:
            err_data = resp.json()
        except ValueError:
            err_data = {"error": resp.text}
        return {"ok": False, "error": err_data.get("error", resp.text)}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": "Could not reach premium evaluation server."}


def _get_bulk_job_status(tid: str, job_id: str) -> dict:
    try:
        resp = requests.get(
            f"{SERVER_URL}/templates/{tid}/calculate-bulk/{job_id}/status",
            timeout=30,
        )
        if resp.ok:
            return {"ok": True, "status": resp.json()}
        try:
            err_data = resp.json()
        except ValueError:
            err_data = {"error": resp.text}
        return {"ok": False, "error": err_data.get("error", resp.text)}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": "Could not reach premium evaluation server."}


def _download_bulk_job_result(tid: str, job_id: str) -> dict:
    try:
        resp = requests.get(
            f"{SERVER_URL}/templates/{tid}/calculate-bulk/{job_id}/download",
            timeout=60,
        )
        if resp.ok:
            return {"ok": True, "content": resp.content}
        try:
            err_data = resp.json()
        except ValueError:
            err_data = {"error": resp.text}
        return {"ok": False, "error": err_data.get("error", resp.text)}
    except requests.exceptions.ConnectionError:
        return {"ok": False, "error": "Could not reach premium evaluation server."}


def _poll_bulk_job_status(tid: str, job_id: str, interval: int = 2, timeout: int = 900) -> dict:
    started = time.time()
    while time.time() - started < timeout:
        status_res = _get_bulk_job_status(tid, job_id)
        if not status_res["ok"]:
            return status_res
        data = status_res["status"]
        if data.get("status") in {"complete", "failed"}:
            return {"ok": True, "status": data}
        time.sleep(interval)
    return {"ok": False, "error": "Bulk job polling timed out after waiting for the job to complete."}


# ── Premium calculation dashboard ──────────────────────────────────────────────

def _fmt_currency(val) -> str:
    try:
        return f"₹ {float(val):,.2f}"
    except (TypeError, ValueError):
        return "—"


def _render_premium_comparison(local_base, local_total, partner_base, partner_total, rounding_rule: str = "nearest") -> None:
    """Premium Comparison (Sanity Check): Base Premium then Total Premium
    (Including GST), always in that order. Computation lives in
    dim_parser.py (_build_comparison_row / _overall_premium_status) so
    Bulk Premium's row loop reuses the exact same logic — this function is
    purely the Streamlit rendering of that shared computation.
    rounding_rule comes from the template's own saved setting (Template
    Builder) — see _build_comparison_row for what it does."""
    rows = [
        _build_comparison_row("Base Premium", local_base, partner_base, rounding_rule),
        _build_comparison_row("Total Premium (Including GST)", local_total, partner_total, rounding_rule),
    ]
    overall_icon, overall_text = _overall_premium_status(rows)

    st.write("")
    st.markdown("#### Premium Comparison (Sanity Check)")
    st.markdown(f"**{overall_icon} {overall_text}**")

    with st.container(border=True):
        h1, h2, h3, h4, h5 = st.columns([2.2, 1.6, 1.6, 1.6, 1.6])
        for col, label in zip((h1, h2, h3, h4, h5), ["Metric", "Local Excel", "Partner API", "Difference", "Status"]):
            with col:
                st.caption(label)

        for row in rows:
            c1, c2, c3, c4, c5 = st.columns([2.2, 1.6, 1.6, 1.6, 1.6])
            with c1:
                st.markdown(f"**{row['label']}**")
            with c2:
                st.markdown(_fmt_currency(row["local"]) if row["local"] is not None else "—")
            with c3:
                st.markdown(_fmt_currency(row["partner"]) if row["partner"] is not None else "—")
            with c4:
                st.markdown(_fmt_currency(row["diff"]) if row["diff"] is not None else "—")
            with c5:
                st.markdown(row["status_label"])


def _render_premium_metrics(data: dict, title: str = "Calculated Premiums") -> None:
    if not data:
        return

    base = data.get("base_premium")
    total = data.get("total_premium")
    if base is None and total is None:
        return

    if base is None:
        base = 0.0
    if total is None:
        total = base * 1.18

    session_token = data.get("session_token")
    api_status = data.get("status") or data.get("result") or "unknown"

    st.write("")
    st.markdown(f"#### {title}")
    with st.container(border=True):
        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            st.metric("Base Premium", _fmt_currency(base))
        with mc2:
            st.metric("Total Premium (Inc. GST)", _fmt_currency(total))
        with mc3:
            if session_token:
                st.metric("Session Token", session_token)
            else:
                st.metric("API Response", api_status)

# ── Matrix Unpivoter Tab ──────────────────────────────────────────────────────

def on_preset_change():
    preset = st.session_state.cfg_si_cond
    if preset == "SI_G_20 (> 20 Lakhs)":
        st.session_state.cfg_si_prefix = "SI_G_20"
        st.session_state.cfg_min_sa = 2000001
        st.session_state.cfg_max_sa = 150000000
    elif preset == "SI_LT_20 (<= 20 Lakhs)":
        st.session_state.cfg_si_prefix = "SI_LT_20"
        st.session_state.cfg_min_sa = 1000
        st.session_state.cfg_max_sa = 2000000
    else:  # CUSTOM
        st.session_state.cfg_si_prefix = "CUSTOM_PREFIX"
        st.session_state.cfg_min_sa = 1000
        st.session_state.cfg_max_sa = 100000000


def render_tab_matrix_unpivoter():
    st.markdown("### 📑 Bajaj Excel Matrix Unpivoter & CSV Generator")
    st.caption("Upload a 2D matrix rate table (Age vs Tenure) and convert it into standard flattened CSV rows.")

    uploaded_matrix = st.file_uploader("Upload Bajaj Rate Matrix (.xlsx)", type=["xlsx"], key="matrix_unpivoter_file")

    st.markdown("#### ⚙️ Configuration Inputs (System Metadata)")
    st.info("Fill in metadata required for the output CSV schema.", icon="ℹ️")

    if "cfg_si_prefix" not in st.session_state:
        st.session_state.cfg_si_prefix = "SI_G_20"
        st.session_state.cfg_min_sa = 2000001
        st.session_state.cfg_max_sa = 150000000

    col1, col2, col3 = st.columns(3)
    
    with col1:
        insurance_code = st.text_input("Insurance Code", value="BAJAJ_GCCP", key="cfg_ins_code")
        cover_code = st.text_input("Cover Code", value="GCCP_RED_COVER", key="cfg_cover_code")
        partner_code = st.text_input("Partner Code", value="GRIHUM", key="cfg_partner_code")
        product_type = st.text_input("Product Type", value="GROUP_CREDIT_PROTECTION", key="cfg_prod_type")

    with col2:
        si_condition = st.selectbox(
            "Sum Insured Band Preset",
            ["SI_G_20 (> 20 Lakhs)", "SI_LT_20 (<= 20 Lakhs)", "CUSTOM"],
            key="cfg_si_cond",
            on_change=on_preset_change
        )

        si_prefix = st.text_input("Code Prefix / Band Slug", key="cfg_si_prefix")
        min_sa = st.number_input("Min Sum Insured (SA)", step=1000, key="cfg_min_sa")
        max_sa = st.number_input("Max Sum Insured (SA)", step=10000, key="cfg_max_sa")

    with col3:
        term_unit = st.selectbox("Term Unit", ["year", "month"], key="cfg_term_unit")
        premium_paying_type = st.selectbox("Premium Paying Type", ["SINGLE", "REGULAR", "LIMITED"], key="cfg_ppt_type")
        gender = st.selectbox("Gender Filter", ["MALE", "FEMALE", "ALL (Leave Blank)"], key="cfg_gender")
        smoker = st.selectbox("Smoker Filter", ["NON_SMOKER", "SMOKER", "ALL (Leave Blank)"], key="cfg_smoker")
        active_status = st.checkbox("Active Status", value=True, key="cfg_active")

    if uploaded_matrix and st.button("Unpivot Matrix & Export CSV", type="primary"):
        try:
            xls = pd.ExcelFile(uploaded_matrix)
            sheet_names = xls.sheet_names
            
            detected_partner = partner_code
            if "Partner Details" in sheet_names:
                df_p = pd.read_excel(xls, "Partner Details")
                if "Partner Code" in df_p.columns:
                    detected_partner = str(df_p.columns[1]).strip()
            
            df_raw = pd.read_excel(xls, sheet_name=0)

            header_row_idx = None
            for idx, row in df_raw.iterrows():
                row_str = row.astype(str).str.cat(sep=' ')
                if "Age" in row_str and "Term" in row_str:
                    header_row_idx = idx
                    break

            if header_row_idx is None:
                st.error("Could not locate the 'Age / Term' header row in sheet.")
                return

            matrix_data = df_raw.iloc[header_row_idx+1:].copy()
            tenure_cols = [c for c in df_raw.iloc[header_row_idx].values[1:] if pd.notna(c)]
            
            # Map gender and smoker values correctly
            selected_gender = "" if gender.startswith("ALL") else gender
            selected_smoker = "" if smoker.startswith("ALL") else smoker

            flattened_rows = []
            
            for _, row in matrix_data.iterrows():
                age_val = row.iloc[0]
                if pd.isna(age_val):
                    continue
                try:
                    age_val = int(age_val)
                except (ValueError, TypeError):
                    continue

                for i, tenure in enumerate(tenure_cols):
                    rate_val = row.iloc[i+1]
                    if pd.isna(rate_val):
                        continue

                    try:
                        tenure_val = float(tenure)
                        rate_float = float(rate_val)
                    except (ValueError, TypeError):
                        continue
                    
                    row_code = f"{si_prefix}_{int(tenure_val)}_{age_val}"
                    plan_group = f"{insurance_code}_{si_prefix}_{int(tenure_val)}_YEAR"

                    flattened_rows.append({
                        "code": row_code,
                        "insurance_code": insurance_code,
                        "cover_code": cover_code,
                        "partner_code": detected_partner,
                        "product_type": product_type,
                        "min_term": tenure_val,
                        "max_term": tenure_val,
                        "term_unit": term_unit,
                        "min_validity": "",
                        "max_validity": "",
                        "validity_unit": term_unit,
                        "min_age": float(age_val),
                        "max_age": float(age_val),
                        "min_sa": float(min_sa),
                        "max_sa": float(max_sa),
                        "rate": rate_float,
                        "active": active_status,
                        "premium_paying_type": premium_paying_type,
                        "premium_paying_term": tenure_val if premium_paying_type != "SINGLE" else 1.0,
                        "payout_type": "",
                        "payment_frequency": "",
                        "gender": selected_gender,
                        "smoker": selected_smoker,
                        "min_annual_income": "",
                        "max_annual_income": "",
                        "discount_type": "",
                        "discount_value": "",
                        "plan_group_code": plan_group,
                        "plan_group_name": plan_group.replace("_", " "),
                        "plan_group_term_unit": term_unit,
                        "plan_group_term": tenure_val,
                        "plan_group_sum_insured": ""
                    })

            result_df = pd.DataFrame(flattened_rows)
            st.success(f"Successfully processed and unpivoted {len(result_df)} rows!")

            st.dataframe(result_df.head(10))

            csv_buffer = io.BytesIO()
            result_df.to_csv(csv_buffer, index=False)
            
            st.download_button(
                label="⬇️ Download Flattened CSV",
                data=csv_buffer.getvalue(),
                file_name=f"{insurance_code}_{si_prefix}_flattened.csv",
                mime="text/csv",
                type="primary"
            )

        except Exception as e:
            st.error(f"Error during matrix processing: {str(e)}")

def render_tab_calculate():
    st.markdown("### Premium Calculation Workflow")
    st.caption(
        "Calculate a premium for one customer, or upload an Excel file to calculate "
        "premiums for many customers at once — both modes use the exact same "
        "calculation pipeline."
    )

    mode = st.radio(
        "Mode", ["Single Customer", "Bulk Upload"],
        horizontal=True, key="calc_mode_select", label_visibility="collapsed",
    )
    st.write("")

    if mode == "Single Customer":
        _render_single_customer_mode()
    else:
        _render_bulk_upload_mode()


def _render_single_customer_mode():
    if st.session_state.calc_flow_type["single"] is None:
        st.info("Choose which API this run should use before continuing.")
        if st.button("Set Up Workflow", key="start_single_workflow_btn", type="primary"):
            _workflow_setup_dialog("single")
        return  # stop drawing the rest of the page until they answer
    st.caption("Each API creates an authenticated session token. Once initialized, the rater calculates the premium identically for all paths.")
    
    flow_type = st.session_state.calc_flow_type["single"]

    if st.button("🔁 Change Workflow", type="primary", use_container_width=True):
        _workflow_setup_dialog("single")

    # Custom payload status + edit trigger for the selected workflow
    if flow_type in st.session_state.custom_payloads:
        st.markdown("✏️ **Using custom payload**")
    else:
        st.caption("Using default (hardcoded) payload for this workflow.")
    

    templates = _fetch_templates()
    if not templates:
        st.warning(
            "No templates found. Go to Tab 1 to define a product, or start your server.",
            icon="⚠️",
        )
        return

    name_to_id = {t["name"]: t["id"] for t in templates}
    selected_name = st.selectbox("Product Template", list(name_to_id.keys()), key="calc_product_select")
    tid = name_to_id[selected_name]

    full_def = _fetch_template_detail(tid)
    if not full_def:
        st.error("Could not fetch rater layout rules.")
        return

    # Lead Inputs (only render if they're relevant)
    lead_details = None
    if flow_type in ["save_lead", "create_lead"]:
        with st.container(border=True):
            st.markdown("👤 **Lead Creation Parameters**")
            lc1, lc2, lc3 = st.columns(3)
            with lc1:
                lead_name = st.text_input("Name", value="John Doe", key="calc_lead_name")
            with lc2:
                lead_mobile = st.text_input("Mobile Number", value="9876543210", key="calc_lead_mobile")
            with lc3:
                lead_email = st.text_input("Email", value="johndoe@example.com", key="calc_lead_email")
            lead_details = {
                "name": lead_name.strip(),
                "mobile": lead_mobile.strip(),
                "email": lead_email.strip()
            }

    st.markdown("#### Rater Dimensions")
    inputs = _render_calc_inputs(full_def)

    # UI session state reset if flow type is switched
    if st.session_state.session_flow_type != flow_type:
        st.session_state.session_token = None
        st.session_state.session_flow_type = flow_type

    # Session Status Area
    st.write("")
    if st.session_state.session_token:
        st.success(f"🔗 Active Session Token: `{st.session_state.session_token}`")
        if st.button("Reset Session/Tokens", key="reset_session_btn"):
            st.session_state.session_token = None
            st.rerun()
    else:
        st.info("No active session. Initiating this action will generate a new session token first.", icon="🔑")

    st.write("")
    if st.button("Initialize & Calculate Premium", type="primary", key="calc_premium_action_btn"):
        with st.spinner("Executing workflow..."):
            
            init_res = _initialize_api_session(
                flow_type,
                lead_details,
                inputs,
                custom_payload=st.session_state.custom_payloads.get(flow_type),
            )
            if not init_res["ok"]:
                st.error(flow_type == "save_loan" and f"Premium fetch failed: {init_res['error']}" or f"Failed to generate session token: {init_res['error']}")
                return

            data = init_res["data"]
            sent_payload = data.get("payload")
            session_token = init_res.get("session_token") or data.get("session_token")
            if session_token:
                st.session_state.session_token = session_token
                st.session_state.session_flow_type = flow_type

            if data.get("status") == "failed":
                error_message = data.get("errors") or data.get("error") or data.get("raw_fetch_quote", {}).get("quotes", [{}])[0].get("error_message")
                if isinstance(error_message, (dict, list)):
                    error_message = json.dumps(error_message, indent=2)
                st.error(f"Partner quote failed: {error_message}")

            _render_premium_metrics(data, title="Partner API Premiums")

            if sent_payload:
                with st.expander("View sent payload"):
                    st.json(sent_payload)

            calc_res = _calculate_premium_with_token(tid, inputs, init_res.get("session_token"))
            if calc_res["ok"]:
                st.write("")
                st.markdown("#### Local Excel Premiums")
                data = calc_res["data"]
                base = data.get("base_premium", 0.0)
                total = data.get("total_premium") or (base * 1.18)

                with st.container(border=True):
                    mc1, mc2, mc3 = st.columns(3)
                    with mc1:
                        st.metric("Base Premium", _fmt_currency(base))
                    with mc2:
                        st.metric("Total Premium (Inc. GST)", _fmt_currency(total))
                    with mc3:
                        if init_res.get("session_token"):
                            st.metric("Associated Session Token", init_res["session_token"])
                        else:
                            st.metric("Session Token", "N/A")

                partner_base = init_res["data"].get("base_premium")
                partner_total = init_res["data"].get("total_premium")
                _render_premium_comparison(
                    base, total, partner_base, partner_total,
                    rounding_rule=full_def.get("rounding_rule", "nearest"),
                )
            else:
                st.error(f"Premium Calculation failed: {calc_res['error']}")


# ── Calculate Premium: Bulk Upload mode ───────────────────────────────────────
# One uploaded Excel of customer rows in, one Excel of the same rows +
# calculated premiums out, via a single backend request
# (POST /templates/<id>/calculate-bulk). Same calculation pipeline as
# Single Customer mode above — including the selected Workflow Action's
# Partner API call — just run once per row instead of once per click.

def _build_sample_template_bytes(full_def: dict) -> bytes:
    """Builds a sample input workbook — header row = each required
    dimension's display name, one example data row of sample values —
    derived live from the template definition via _template_required_dimensions,
    so it can never drift out of sync with what the template actually needs."""
    required = _template_required_dimensions(full_def)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Input"
    ws.append([d["name"] for d in required])
    ws.append([d["sample"] for d in required])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_sample_json_bytes(full_def: dict) -> bytes:
    """Builds a sample JSON payload — a one-element list whose keys are
    exactly the same required dimension names _build_sample_template_bytes
    uses for the Excel sample, so both stay in sync with the template
    definition. Values are left blank for the user to fill in."""
    required = _template_required_dimensions(full_def)
    sample = [{d["name"]: "" for d in required}]
    return json.dumps(sample, indent=2).encode("utf-8")


def _json_rows_to_excel_bytes(rows: list) -> bytes:
    """Normalizes a list of flat {column_name: value} dicts (parsed from an
    uploaded/pasted JSON payload) into the exact same row/column shape the
    Excel upload path already produces — an in-memory workbook with one
    header row and one row per customer. This is purely a data-shape
    conversion: it doesn't touch, know about, or duplicate any calculation
    logic. Once built, these bytes are handed to the exact same
    /calculate-bulk endpoint an uploaded .xlsx file would be — the
    calculation pipeline never knows whether the input originated as
    Excel or JSON."""
    headers = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                headers.append(key)
                seen.add(key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_json_payload(raw: str):
    """Parses and validates a JSON payload for Bulk Premium. Returns
    (rows, None) on success or (None, error_message) on failure — never
    raises, so the caller can surface a clean st.error either way."""
    raw = (raw or "").strip()
    if not raw:
        return None, "No JSON provided."
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        return None, f"Invalid JSON: {e}"

    if not isinstance(parsed, list):
        return None, "JSON payload must be a list of customer records, e.g. [{...}, {...}]."
    if not parsed:
        return None, "JSON payload is empty — it must contain at least one customer record."
    if not all(isinstance(r, dict) for r in parsed):
        return None, "Every item in the JSON list must be an object, e.g. {\"Age\": 30, ...}."

    return parsed, None


def _spreadsheet_rows_to_excel_bytes(df: "pd.DataFrame", required: list[dict]) -> bytes | None:
    """Converts the edited st.data_editor DataFrame (Spreadsheet input
    mode) into the exact same in-memory Excel bytes shape that the Excel
    Workbook and JSON Payload paths already produce, by reusing
    _json_rows_to_excel_bytes — i.e. this is purely a data-shape
    conversion, never a separate calculation path. Fully blank rows
    (every required column empty) are dropped before conversion so
    trailing/unused starter rows don't get sent to the pipeline. Returns
    None if no non-blank rows remain."""
    if df is None or df.empty:
        return None

    required_cols = [d["name"] for d in required]

    rows = []
    for _, series in df.iterrows():
        row = {}
        has_value = False
        for col in required_cols:
            val = series.get(col, "")
            if pd.isna(val):
                val = ""
            val = val if isinstance(val, str) else val
            if isinstance(val, str) and val.strip() == "":
                val = ""
            else:
                has_value = True
            row[col] = val
        if has_value:
            rows.append(row)

    if not rows:
        return None

    return _json_rows_to_excel_bytes(rows)


def _render_bulk_upload_mode():
    if st.session_state.calc_flow_type["bulk"] is None:
        st.info("Choose which API this run should use before continuing.")
        if st.button("Set Up Workflow", key="start_bulk_workflow_btn", type="primary"):
            _workflow_setup_dialog("bulk")
        return  # stop drawing the rest of the page until they answer
    st.caption(
        "Upload an Excel file with one row per customer to calculate premiums for all of "
        "them at once, using a single selected template — the same calculation pipeline as "
        "Single Customer mode, just run for every row in one batch instead of one "
        "customer at a time."
    )

    templates = _fetch_templates()
    if not templates:
        st.warning(
            "No templates found. Go to Tab 1 to define a product, or start your server.",
            icon="⚠️",
        )
        return

    flow_type = st.session_state.calc_flow_type["bulk"]

    if st.button("🔁 Change Workflow", type="primary", use_container_width=True):
        _workflow_setup_dialog("bulk")

    # Custom payload status + edit trigger — shared session state with
    # Single Customer mode, so setting it once (in either mode) applies to
    # both, and it's applied identically to every row in this bulk run.
    if flow_type in st.session_state.custom_payloads:
        st.markdown("✏️ **Using custom payload**")
    else:
        st.caption("Using default (hardcoded) payload for this workflow.")

    flow_type_labels = {"save_loan": "Save Loan", "save_lead": "Save Lead", "create_lead": "Create Lead"}
    if flow_type in ("save_lead", "create_lead"):
        st.warning(
            f"**{flow_type_labels[flow_type]}** creates a real lead at the partner institution for "
            f"every row. A file with N customer rows will create N real partner-side leads "
            f"— there is no dry-run for this option.",
            icon="⚠️",
        )
    name_to_id = {t["name"]: t["id"] for t in templates}
    selected_name = st.selectbox("Product Template", list(name_to_id.keys()), key="bulk_product_select")
    tid = name_to_id[selected_name]

    full_def = _fetch_template_detail(tid)
    if not full_def:
        st.error("Could not fetch rater layout rules.")
        return

    required = _template_required_dimensions(full_def)

    st.write("")
    st.markdown("**Required columns for this template:**")
    st.caption(", ".join(d["name"] for d in required) if required else "—")

    st.write("")
    input_source = st.radio(
        "Input Source", ["Excel Workbook", "JSON Payload", "Spreadsheet"],
        horizontal=True, key="bulk_input_source",
    )

    bulk_input_bytes = None
    bulk_input_filename = "input.xlsx"

    if input_source == "Excel Workbook":
        # Unchanged from before — same sample download, same uploader, same bytes.
        col_sample, _ = st.columns([2, 3])
        with col_sample:
            sample_bytes = _build_sample_template_bytes(full_def)
            st.download_button(
                "⬇️ Download Sample Template",
                data=sample_bytes,
                file_name=f"sample_input_{selected_name.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.write("")
        uploaded = st.file_uploader("Upload customer records (.xlsx)", type=["xlsx"], key="bulk_upload_file")
        if uploaded is not None:
            bulk_input_bytes = uploaded.getvalue()
            bulk_input_filename = uploaded.name

    elif input_source == "JSON Payload":
        col_sample, _ = st.columns([2, 3])
        with col_sample:
            sample_json_bytes = _build_sample_json_bytes(full_def)
            st.download_button(
                "⬇️ Download Sample JSON",
                data=sample_json_bytes,
                file_name=f"sample_input_{selected_name.replace(' ', '_')}.json",
                mime="application/json",
                use_container_width=True,
            )

        st.write("")
        json_file = st.file_uploader("Upload customer records (.json)", type=["json"], key="bulk_upload_json_file")
        json_text = st.text_area(
            "...or paste JSON here", key="bulk_json_text", height=180,
            placeholder='[\n  {"Age": 30, "Loan Type": "HL"},\n  {"Age": 45, "Loan Type": "LAP"}\n]',
        )

        raw_json = json_file.getvalue().decode("utf-8") if json_file is not None else json_text
        if raw_json and raw_json.strip():
            rows, err = _parse_json_payload(raw_json)
            if err:
                st.error(err, icon="⚠️")
            else:
                bulk_input_bytes = _json_rows_to_excel_bytes(rows)
                bulk_input_filename = "input_from_json.xlsx"

    else:  # "Spreadsheet"
        st.caption(
            "Fill in customer records directly below — edit cells, add rows, or delete "
            "rows, just like an Excel sheet. Columns are generated automatically from "
            "the selected template's required dimensions."
        )

        required_cols = [d["name"] for d in required]

        # Reset initial data ONLY if template required columns change
        cols_key = "bulk_spreadsheet_cols"
        initial_data_key = "bulk_spreadsheet_initial"
        
        if cols_key not in st.session_state or set(st.session_state.get(cols_key) or []) != set(required_cols):
            st.session_state[cols_key] = required_cols
            st.session_state[initial_data_key] = pd.DataFrame(
                [{c: "" for c in required_cols} for _ in range(5)],
                columns=required_cols,
            )
            # Clear previous editor widget state if template changes
            if "bulk_spreadsheet_editor" in st.session_state:
                del st.session_state["bulk_spreadsheet_editor"]

        edited_df = st.data_editor(
            st.session_state[initial_data_key],
            num_rows="dynamic",
            use_container_width=True,
            key="bulk_spreadsheet_editor",
        )

        bulk_input_bytes = _spreadsheet_rows_to_excel_bytes(edited_df, required)
        bulk_input_filename = "input_from_spreadsheet.xlsx"

    st.write("")
    if st.button("Calculate Bulk Premiums", type="primary", key="bulk_calc_btn", disabled=bulk_input_bytes is None):
            with st.spinner("Starting bulk job..."):
                start_res = _start_bulk_premium_job(
                    tid, bulk_input_bytes, bulk_input_filename, flow_type,
                    custom_payload=st.session_state.custom_payloads.get(flow_type),
                )

            if not start_res["ok"]:
                st.error(start_res["error"], icon="⚠️")
            else:
                job_id = start_res["job_id"]
                with st.spinner("Bulk job running — polling status..."):
                    status_res = _poll_bulk_job_status(tid, job_id, interval=2, timeout=900)

                if not status_res["ok"]:
                    st.error(status_res["error"], icon="⚠️")
                else:
                    job_status = status_res["status"]
                    if job_status["status"] == "failed":
                        st.error(job_status.get("error_message") or "Bulk job failed.", icon="⚠️")
                    elif job_status["status"] == "complete":
                        download_res = _download_bulk_job_result(tid, job_id)
                        if not download_res["ok"]:
                            st.error(download_res["error"], icon="⚠️")
                        else:
                            st.session_state["bulk_result_bytes"] = download_res["content"]
                            st.session_state["bulk_result_name"] = job_status.get("result_filename") or f"bulk_premium_results_{selected_name.replace(' ', '_')}.xlsx"
    if st.session_state.get("bulk_result_bytes"):
        out_wb = openpyxl.load_workbook(io.BytesIO(st.session_state["bulk_result_bytes"]))
        out_ws = out_wb.active
        rows = list(out_ws.iter_rows(min_row=2, values_only=True))
        header = [c.value for c in next(out_ws.iter_rows(min_row=1, max_row=1))]

        def _col(name):
            return header.index(name) if name in header else None

        proc_idx = _col("Processing Status")
        overall_idx = _col("Overall Comparison")

        total = len(rows)
        ok_count = sum(1 for r in rows if proc_idx is not None and r[proc_idx] == "OK") if proc_idx is not None else total
        error_count = total - ok_count
        mismatch_count = sum(
            1 for r in rows if overall_idx is not None and r[overall_idx] and "Mismatch" in str(r[overall_idx])
        )

        st.write("")
        overview_cols = st.columns(4)
        overview_cols[0].metric("Total rows", total)
        overview_cols[1].metric("Successful rows", ok_count)
        overview_cols[2].metric("Error rows", error_count)
        overview_cols[3].metric("Mismatch rows", mismatch_count)

        summary = f"{ok_count}/{total} rows processed successfully"
        if mismatch_count:
            summary += f", {mismatch_count} premium mismatch(es) detected"
        if error_count:
            summary += f", {error_count} row(s) had processing errors"
        (st.warning if (error_count or mismatch_count) else st.success)(summary, icon="⚠️" if (error_count or mismatch_count) else "✅")

        with st.expander("View row-level result details", expanded=False):
            preview_headers = [
                h for h in [
                    "Processing Status",
                    "Overall Comparison",
                    "Partner API Status",
                    "Partner API Error",
                    "Error Message",
                ] if h in header
            ]
            if not preview_headers:
                preview_headers = header[: min(8, len(header)) ]

            preview_rows = []
            for r in rows[:min(20, len(rows))]:
                preview_rows.append({header[i]: r[i] for i in range(len(header)) if header[i] in preview_headers})

            if preview_rows:
                st.dataframe(pd.DataFrame(preview_rows).fillna(""), use_container_width=True)
            else:
                st.write("No preview data available.")

        st.download_button(
            "⬇️ Download Results",
            data=st.session_state["bulk_result_bytes"],
            file_name=st.session_state["bulk_result_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def main():
    init_state()

    st.title("Offline rater automation")
    st.write("")

    tab0, tab1, tab2, tab3, tab4 = st.tabs([
        "Dimension Library",
        "Template Builder",
        "Upload & Flatten",
        "Calculate Premium",
        "Matrix Unpivoter (CSV)"
    ])

    with tab0:
        render_tab_dimension_library()

    with tab1:
        render_tab_template_builder()

    with tab2:
        render_tab_upload()

    with tab3:
        render_tab_calculate()

    with tab4:
        render_tab_matrix_unpivoter()


if __name__ == "__main__":
    main()