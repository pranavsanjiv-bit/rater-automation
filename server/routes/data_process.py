# pyrefly: ignore [missing-import]
from flask import Blueprint, jsonify, request   

from frontend.dim_parser import get_dimension_slug
from server.db import get_db
from server.helpers import _read_workbook_meta, get_value_slug, _normalize_float_to_str
import openpyxl, json, uuid
from datetime import datetime, timezone


processing_bp = Blueprint('processing', __name__)

@processing_bp.post("/flatten")
def flatten():
    if openpyxl is None:
        return jsonify({"error": "openpyxl is not installed on the server."}), 500

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded."}), 400

    wb = openpyxl.load_workbook(file, data_only=True)

    try:
        meta = _read_workbook_meta(wb)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    template_id = request.form.get("template_id") or meta.get("template_id")
    if not template_id:
        return jsonify({
            "error": "No template_id in form data and none embedded in workbook _meta sheet.",
        }), 400

    sheet_mapping = meta.get("sheet_mapping", {})

    with get_db() as conn:
        row = conn.execute(
            "SELECT definition_json FROM rate_templates WHERE id = ?", (template_id,)
        ).fetchone()
    if not row:
        return jsonify({"error": f"Template '{template_id}' not found."}), 404

    full_def = json.loads(row["definition_json"])
    definition = full_def["definition"]
    axes = definition.get("axes", {})
    row_axis = axes.get("row")
    col_axis = axes.get("col")
    two_d = col_axis is not None

    # Get dimension slugs for row/col axes (for dims_debug)
    row_dim_slug = get_dimension_slug(row_axis["name"]) if row_axis else None
    col_dim_slug = get_dimension_slug(col_axis["name"]) if col_axis else None

    rows_to_upsert = []
    blank_cells = []
    now = datetime.now(timezone.utc).isoformat()

    skipped_sheets = {"Index", "_meta"}

    for ws in wb.worksheets:
        if ws.title in skipped_sheets:
            continue

        sheet_name = ws.title
        outer_dim_slugs = sheet_mapping.get(sheet_name)
        if outer_dim_slugs is None:
            continue

        if two_d:
            # Row 1 = column headers (col axis values), Col A = row axis values
            col_headers = {}  # col_index -> col_axis_val_slug
            for cell in ws[1]:
                if cell.column == 1:
                    continue
                if cell.value is not None:
                    col_headers[cell.column] = get_value_slug(_normalize_float_to_str(cell.value))

            for row in ws.iter_rows(min_row=2):
                row_val_cell = row[0]
                if row_val_cell.value is None:
                    continue
                row_val_slug = get_value_slug(_normalize_float_to_str(row_val_cell.value))

                for cell in row[1:]:
                    if cell.column not in col_headers:
                        continue
                    col_val_slug = col_headers[cell.column]

                    # Build lookup key
                    outer_parts = [
                        f"{ds}_{outer_dim_slugs[ds]}"
                        for ds in sorted(outer_dim_slugs)
                    ]
                    key = "-".join(outer_parts + [row_val_slug, col_val_slug])
                    dims_debug = {**outer_dim_slugs}
                    if row_dim_slug:
                        dims_debug[row_dim_slug] = row_val_slug
                    if col_dim_slug:
                        dims_debug[col_dim_slug] = col_val_slug

                    if cell.value is None:
                        blank_cells.append({
                            "sheet": ws.title,
                            "row": cell.row,
                            "col": cell.column,
                            "key": key,
                        })
                        continue

                    try:
                        val = float(cell.value)
                    except (TypeError, ValueError):
                        blank_cells.append({"sheet": ws.title, "row": cell.row, "col": cell.column, "key": key})
                        continue

                    rows_to_upsert.append((template_id, key, json.dumps(dims_debug), val, now))
        else:
            # 1D: col A = axis label, col B = value; row 1 is header
            for row in ws.iter_rows(min_row=2):
                if len(row) < 2:
                    continue
                axis_cell, val_cell = row[0], row[1]
                if axis_cell.value is None:
                    continue
                axis_val_slug = get_value_slug(_normalize_float_to_str(axis_cell.value))

                outer_parts = [
                    f"{ds}_{outer_dim_slugs[ds]}"
                    for ds in sorted(outer_dim_slugs)
                ]
                key = "-".join(outer_parts + [axis_val_slug])
                dims_debug = {**outer_dim_slugs}
                if row_dim_slug:
                    dims_debug[row_dim_slug] = axis_val_slug

                if val_cell.value is None:
                    blank_cells.append({"sheet": ws.title, "row": val_cell.row, "col": 2, "key": key})
                    continue

                try:
                    val = float(val_cell.value)
                except (TypeError, ValueError):
                    blank_cells.append({"sheet": ws.title, "row": val_cell.row, "col": 2, "key": key})
                    continue
                    
                rows_to_upsert.append((template_id, key, json.dumps(dims_debug), val, now))

    with get_db() as conn:
        conn.executemany(
            "INSERT INTO rate_values (template_id, lookup_key, dims_json, value, updated_at) "
            "VALUES (?, ?, ?, ?, ?) "
            "ON CONFLICT(template_id, lookup_key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at",
            rows_to_upsert,
        )
    cells_written = len(rows_to_upsert)
    total = cells_written + len(blank_cells)
    pct = round(cells_written / total * 100, 1) if total else 0

    return jsonify({
        "template_id": template_id,
        "cells_written": cells_written,
        "blank_cells_count": len(blank_cells),
        "completion_pct": pct,
        "blank_cells": blank_cells[:50],  # cap to avoid huge responses
    })

