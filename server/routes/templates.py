# ==========================================================
# Imports
# ==========================================================

import json
import re
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone, date
from io import BytesIO

# pyrefly: ignore [missing-import]
from flask import Blueprint, jsonify, request, send_file
import openpyxl

from frontend.dim_parser import (
    get_dimension_slug,
    _template_required_dimensions,
    _build_comparison_row,
    _overall_premium_status,
)
from server.db import get_db
from server.helpers import (
    _parse_template_body,
    _calculate_premium_for_inputs,
)
from server.routes.injestion import (
    _run_partner_calculation,
    _run_save_lead_calculation,
    _run_create_lead_calculation,
)

_bulk_job_executor = ThreadPoolExecutor(max_workers=2)
_bulk_job_store: dict[str, dict] = {}
_bulk_job_store_lock = threading.Lock()

templates_bp = Blueprint('templates', __name__, url_prefix='/templates')


def _get_bulk_job(job_id: str) -> dict | None:
    with _bulk_job_store_lock:
        return _bulk_job_store.get(job_id)


def _update_bulk_job(job_id: str, **updates) -> None:
    with _bulk_job_store_lock:
        job = _bulk_job_store.get(job_id)
        if not job:
            return
        job.update(updates)
        job["updated_at"] = datetime.now(timezone.utc).isoformat()


def _run_bulk_job(job_id: str, template_id: str, full_def: dict, file_bytes: bytes, flow_type: str, direct_payload: dict | None):
    _update_bulk_job(job_id, status="running", error=None, error_message=None)
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1), [])
        original_headers = [c.value for c in header_row if c.value is not None]
        if not original_headers:
            raise ValueError("Uploaded file has no header row.")

        header_slugs = {}
        for header in original_headers:
            text = str(header).strip()
            slug = get_dimension_slug(text)
            header_slugs[slug] = header
            alternates = {
                re.sub(r"\s+", "_", text),
                text.replace("-", "_"),
                text.replace("_", " "),
                re.sub(r"[-_\s]+", "_", text),
            }
            for alt in alternates:
                if alt:
                    header_slugs[get_dimension_slug(alt)] = header

        required = _template_required_dimensions(full_def)
        missing = [
            d["name"] for d in required
            if d["slug"] not in header_slugs
            and not (d["slug"] == "age" and "dob" in header_slugs)
        ]
        if missing:
            raise ValueError(f"Uploaded file is missing required column(s): {', '.join(missing)}")

        rows = list(ws.iter_rows(min_row=2))
        total_rows = sum(1 for excel_row in rows if any(c.value is not None for c in excel_row))
        _update_bulk_job(job_id, rows_total=total_rows)

        out_rows = []
        success_rows = 0
        error_rows = 0
        mismatch_rows = 0

        flow_dispatch = {
            "save_loan": lambda dims: _run_partner_calculation("save_loan", dims, direct_payload),
            "save_lead": lambda dims: _run_save_lead_calculation(dims, direct_payload=direct_payload),
            "create_lead": lambda dims: _run_create_lead_calculation(dims, direct_payload=direct_payload),
        }

        for excel_row in rows:
            if all(c.value is None for c in excel_row):
                continue

            row_values = {
                original_headers[i]: excel_row[i].value if i < len(excel_row) else None
                for i in range(len(original_headers))
            }

            inputs = {}
            for slug, header in header_slugs.items():
                val = row_values.get(header)
                if isinstance(val, (datetime, date)):
                    val = val.strftime("%d-%m-%Y")
                inputs[slug] = val

            out_row = dict(row_values)
            errors = []

            local_result, local_status = _calculate_premium_for_inputs(
                full_def, template_id, dict(inputs), conn=None
            )
            local_base = local_result.get("base_premium") if local_status == 200 else None
            local_total = None
            if local_status == 200:
                local_total = local_result.get("total_premium") or (
                    local_base * 1.18 if isinstance(local_base, (int, float)) else None
                )
            else:
                errors.append(f"Local calculation: {local_result.get('error', 'Unknown error')}")

            partner_base = partner_total = None
            partner_status = None
            partner_error_detail = None
            try:
                partner_result = flow_dispatch[flow_type](dict(inputs))
                partner_base = partner_result.get("base_premium")
                partner_total = partner_result.get("total_premium")
                partner_status = partner_result.get("status") or partner_result.get("result")
                partner_error_detail = _extract_partner_error_detail(partner_result)

                if partner_total is None and partner_base is not None:
                    try:
                        partner_total = round(float(partner_base) * 1.18, 2)
                    except (TypeError, ValueError):
                        pass
                if partner_base is None and partner_total is not None:
                    try:
                        partner_base = round(float(partner_total) / 1.18, 2)
                    except (TypeError, ValueError):
                        pass

                if partner_base is None and partner_total is None:
                    if partner_error_detail is not None:
                        errors.append(f"Partner API: {partner_error_detail}")
                    else:
                        errors.append(
                            f"Partner API: premium result missing from response"
                            f" (status={partner_status})"
                        )
            except Exception as e:
                partner_error_detail = str(e)
                errors.append(f"Partner API: {partner_error_detail}")

            rounding_rule = full_def.get("rounding_rule", "nearest")
            comparison_rows = [
                _build_comparison_row("Base Premium", local_base, partner_base, rounding_rule),
                _build_comparison_row("Total Premium (Including GST)", local_total, partner_total, rounding_rule),
            ]
            overall_icon, overall_text = _overall_premium_status(comparison_rows)

            base_row, total_row = comparison_rows
            out_row["Base Premium (Local Excel)"] = base_row["local"]
            out_row["Base Premium (Partner API)"] = base_row["partner"]
            out_row["Base Premium Diff"] = base_row["diff"]
            out_row["Base Premium Status"] = base_row["status_label"]
            out_row["Total Premium (Local Excel)"] = total_row["local"]
            out_row["Total Premium (Partner API)"] = total_row["partner"]
            out_row["Total Premium Diff"] = total_row["diff"]
            out_row["Total Premium Status"] = total_row["status_label"]
            out_row["Partner API Status"] = partner_status
            out_row["Partner API Error"] = partner_error_detail
            out_row["Overall Comparison"] = f"{overall_icon} {overall_text}"
            processing_ok = local_status == 200 and len(errors) == 0
            out_row["Processing Status"] = "OK" if processing_ok else "Error"
            out_row["Error Message"] = "; ".join(errors)

            out_rows.append(out_row)
            if processing_ok:
                success_rows += 1
            else:
                error_rows += 1
            if overall_text.lower().strip().startswith("mismatch"):
                mismatch_rows += 1

            _update_bulk_job(
                job_id,
                rows_processed=len(out_rows),
                success_rows=success_rows,
                error_rows=error_rows,
                mismatch_rows=mismatch_rows,
            )

        out_headers = original_headers + [
            "Base Premium (Local Excel)", "Base Premium (Partner API)", "Base Premium Diff", "Base Premium Status",
            "Total Premium (Local Excel)", "Total Premium (Partner API)", "Total Premium Diff", "Total Premium Status",
            "Partner API Status", "Partner API Error",
            "Overall Comparison", "Processing Status", "Error Message",
        ]

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Bulk Premium Results"
        out_ws.append(out_headers)
        for r in out_rows:
            out_ws.append([r.get(h) for h in out_headers])

        buf = BytesIO()
        out_wb.save(buf)
        result_bytes = buf.getvalue()

        _update_bulk_job(
            job_id,
            status="complete",
            result_bytes=result_bytes,
            result_filename=f"bulk_premium_results_{template_id[:8]}.xlsx",
            download_url=f"/templates/{template_id}/calculate-bulk/{job_id}/download",
        )
    except Exception as exc:
        _update_bulk_job(job_id, status="failed", error=str(exc), error_message=str(exc))


@templates_bp.post("/<template_id>/calculate-bulk")
def calculate_bulk(template_id: str):
    with get_db() as conn:
        row = conn.execute(
            "SELECT definition_json FROM rate_templates WHERE id = ?", (template_id,)
        ).fetchone()
    if not row:
        return jsonify({"error": f"Template '{template_id}' not found."}), 404

    full_def = json.loads(row["definition_json"])

    flow_type = request.form.get("flow_type", "").strip()

    direct_payload = None
    raw_payload = request.form.get("payload", "").strip()
    if raw_payload:
        try:
            parsed_payload = json.loads(raw_payload)
        except json.JSONDecodeError as e:
            return jsonify({"error": f"Invalid custom payload JSON: {e}"}), 400
        if not isinstance(parsed_payload, dict):
            return jsonify({"error": "Custom payload must be a JSON object."}), 400
        direct_payload = parsed_payload

    if flow_type not in {"save_loan", "save_lead", "create_lead"}:
        return jsonify({
            "error": f"Invalid or missing flow_type '{flow_type}'. Bulk Premium supports exactly the same flows as Calculate Premium: 'save_loan', 'save_lead', or 'create_lead'.",
        }), 400

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded."}), 400

    try:
        file_bytes = file.read()
    except Exception as exc:
        return jsonify({"error": f"Could not read uploaded file: {exc}"}), 400

    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    with _bulk_job_store_lock:
        _bulk_job_store[job_id] = {
            "job_id": job_id,
            "template_id": template_id,
            "flow_type": flow_type,
            "status": "pending",
            "created_at": now,
            "updated_at": now,
            "rows_total": 0,
            "rows_processed": 0,
            "success_rows": 0,
            "error_rows": 0,
            "mismatch_rows": 0,
            "result_filename": f"bulk_premium_results_{template_id[:8]}.xlsx",
            "result_bytes": None,
            "download_url": None,
            "error": None,
            "error_message": None,
        }

    _bulk_job_executor.submit(
        _run_bulk_job,
        job_id,
        template_id,
        full_def,
        file_bytes,
        flow_type,
        direct_payload,
    )

    return jsonify({
        "job_id": job_id,
        "status_url": f"/templates/{template_id}/calculate-bulk/{job_id}/status",
    }), 202


@templates_bp.get("/<template_id>/calculate-bulk/<job_id>/status")
def get_bulk_job_status(template_id: str, job_id: str):
    job = _get_bulk_job(job_id)
    if not job or job["template_id"] != template_id:
        return jsonify({"error": "Bulk job not found."}), 404

    response = {
        "job_id": job_id,
        "status": job["status"],
        "rows_total": job["rows_total"],
        "rows_processed": job["rows_processed"],
        "success_rows": job["success_rows"],
        "error_rows": job["error_rows"],
        "mismatch_rows": job["mismatch_rows"],
        "created_at": job["created_at"],
        "updated_at": job["updated_at"],
        "error": job["error"],
        "error_message": job["error_message"],
        "download_url": job["download_url"] if job["status"] == "complete" else None,
    }
    return jsonify(response), 200


@templates_bp.get("/<template_id>/calculate-bulk/<job_id>/download")
def download_bulk_job_result(template_id: str, job_id: str):
    job = _get_bulk_job(job_id)
    if not job or job["template_id"] != template_id:
        return jsonify({"error": "Bulk job not found."}), 404
    if job["status"] != "complete":
        return jsonify({"error": "Bulk job result is not yet available."}), 409
    if not job.get("result_bytes"):
        return jsonify({"error": "Bulk job completed without result data."}), 500

    buf = BytesIO(job["result_bytes"])
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=job["result_filename"],
    )


# ==========================================================
# Template API Routes
# ==========================================================
def _extract_partner_error_detail(response):
    if response is None:
        return None
    if isinstance(response, str):
        return response.strip() or None

    if isinstance(response, dict):
        for key in ("errors", "error", "error_message", "message", "messages", "message_body"):
            value = response.get(key)
            if value not in (None, ""):
                if isinstance(value, list):
                    return "; ".join(str(v) for v in value)
                return value

        if "reason" in response:
            value = response.get("reason")
            if value not in (None, ""):
                if isinstance(value, list):
                    return "; ".join(str(v) for v in value)
                return value

        if "raw_save_loan_response" in response and isinstance(response["raw_save_loan_response"], dict):
            nested = _extract_partner_error_detail(response["raw_save_loan_response"])
            if nested:
                return nested

        if "quotes" in response:
            quotes = response.get("quotes")
            if isinstance(quotes, dict):
                quotes = [quotes]
            if isinstance(quotes, list):
                for quote in quotes:
                    if isinstance(quote, dict):
                        nested = _extract_partner_error_detail(quote)
                        if nested:
                            return nested

        if "data" in response and isinstance(response["data"], dict):
            nested = _extract_partner_error_detail(response["data"])
            if nested:
                return nested

    return None


def _format_json_for_excel(value):
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        except TypeError:
            return str(value)
    return str(value)


# ==========================================================
# Template API Routes
# ==========================================================
# Create Template
@templates_bp.post("/")
def create_template():
    """Create and store a new rate template."""
    body = request.get_json(force=True)
    if not body:
        return jsonify({"error": "Empty request body."}), 400

    full_def, name, content_hash = _parse_template_body(body)
    if full_def is None:
        return jsonify({"error": content_hash}), 400

    template_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    with get_db() as conn:
        conn.execute(
            "INSERT INTO rate_templates (id, name, definition_json, content_hash, created_at) "
            "VALUES (?, ?, ?, ?, ?)",
            (template_id, name, json.dumps(full_def), content_hash, now),
        )

    return jsonify({"template_id": template_id, "name": name}), 201


# ── PUT /templates/<id> ─────────────────────────────────────────────────────
@templates_bp.put("/<template_id>")
def update_template(template_id: str):
    """Update an existing rate template."""
    body = request.get_json(force=True)
    if not body:
        return jsonify({"error": "Empty request body."}), 400

    full_def, name, content_hash = _parse_template_body(body)
    if full_def is None:
        return jsonify({"error": content_hash}), 400

    with get_db() as conn:
        row = conn.execute(
            "SELECT id FROM rate_templates WHERE id = ?", (template_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": f"Template '{template_id}' not found."}), 404

        conn.execute(
            "UPDATE rate_templates SET name = ?, definition_json = ?, content_hash = ? WHERE id = ?",
            (name, json.dumps(full_def), content_hash, template_id),
        )

    return jsonify({"template_id": template_id, "name": name}), 200


# ── GET /templates ────────────────────────────────────────────────────────────

@templates_bp.get("")
def list_templates():
    """Return all available rate templates."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, name, created_at FROM rate_templates ORDER BY created_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


# ── GET /templates/<id> ───────────────────────────────────────────────────────

@templates_bp.get("/<template_id>")
def get_template(template_id: str):
    """Fetch a rate template by ID."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, name, definition_json, created_at FROM rate_templates WHERE id = ?",
            (template_id,),
        ).fetchone()
    if not row:
        return jsonify({"error": f"Template '{template_id}' not found."}), 404

    full_def = json.loads(row["definition_json"])
    return jsonify({
        "id": row["id"],
        "name": row["name"],
        "created_at": row["created_at"],
        **full_def,
    })


# ── POST /templates/<id>/calculate ───────────────────────────────────────────

@templates_bp.post("/<template_id>/calculate")
def calculate(template_id: str):
    """Calculate the premium using the selected template."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT definition_json FROM rate_templates WHERE id = ?", (template_id,)
        ).fetchone()
    if not row:
        return jsonify({"error": f"Template '{template_id}' not found."}), 404

    full_def = json.loads(row["definition_json"])

    body = request.get_json(force=True)
    inputs = body.get("inputs", {})

    result, status = _calculate_premium_for_inputs(full_def, template_id, inputs)
    return jsonify(result), status